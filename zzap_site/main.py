import time
import json

import requests

import xml.etree.ElementTree as ET
from openpyxl import load_workbook


def get_regions(api_key: str):
    # URL API
    url = "https://api.zzap.pro/webservice/datasharing.asmx/GetRegionsV2"

    # Формируем параметры запроса
    params = {
        'login': '',
        'password': '',
        'api_key': api_key,
    }

    try:
        # Выполняем POST-запрос
        response = requests.get(url=url, data=params, timeout=60)

        # Проверяем статус код
        if response.status_code != 200:
            return {"error": f"HTTP error: {response.status_code}"}

        # Преобразуем ответ в JSON
        json_data = response.json()

        # Проверяем на наличие ошибок в ответе
        if json_data.get("error"):
            return {"error": json_data["error"]}

        return json_data  # Возвращаем данные о регионах

    except Exception as e:
        return {"error": str(e)}


# Функция для отправки одного запроса
def get_data_from_api(code: str, brand: str, api_key: str) -> str | None:
    # URL API
    url = "https://api.zzap.pro/webservice/datasharing.asmx/GetSearchResultLight"

    # Параметры запроса
    params = {
        'login': '',
        'password': '',
        'code_region': '1',
        'search_text': '',
        'partnumber': code,
        'class_man': brand,
        'row_count': '100',
        'type_request': '4',
        'api_key': api_key
    }

    try:
        # Отправляем запрос к API
        response = requests.get(url=url, params=params, timeout=60)

        # Проверяем успешность запроса
        if response.status_code == 200:
            return response.text  # Возвращаем текст ответа
        else:
            print(f'Ошибка {response.status_code}: {response.text}')
            return None
    except requests.exceptions.RequestException as e:
        print(f'Ошибка подключения: {e}')
        return None


# Функция для извлечения минимальной цены из XML ответа
def extract_min_price_from_response(xml_data):
    # Парсим XML-ответ
    try:
        root = ET.fromstring(xml_data)

        # Извлекаем JSON-строку
        json_data = root.text  # Текст внутри <string>...</string>

        # Конвертируем в словарь Python
        dict_data = json.loads(json_data)

        min_price = dict_data.get('price_min_instock')

        if min_price:
            return min_price
        return None

    except Exception as e:
        print(f'Ошибка при парсинге JSON: {e}')
        return None


# Основная функция
def process_excel(input_file: str, api_key: str, interval: int = 5):
    # Загружаем рабочую книгу и активный лист
    try:
        workbook = load_workbook(input_file)
        work_sheet = workbook.active  # Доступ к активному листу
    except Exception as e:
        print(f'Ошибка чтения файла: {e}')
        return

    # Индексы столбцов для "Номенклатура.Производитель", "Артикул", "Цена"
    try:
        brand_column = 1
        code_column = 4
        old_price_column = 10
        price_column = 11
    except ValueError as e:
        print(f'Ошибка: не найдены необходимые столбцы. {e}')
        return

    # Обрабатываем строки (начиная с 11-й строки, т.к. 10-я - это заголовки)
    for row_idx in range(12, work_sheet.max_row + 1):
        brand = work_sheet.cell(row=row_idx, column=brand_column).value
        code = work_sheet.cell(row=row_idx, column=code_column).value
        old_price = work_sheet.cell(row=row_idx, column=old_price_column).value

        if brand is None or code is None:
            continue

        # Выполняем запрос к API
        print(f'Отправляем запрос для артикула: {code}, бренд: {brand}')
        xml_data = get_data_from_api(code=code, brand=brand, api_key=api_key)

        if xml_data:
            # Извлекаем минимальную цену из ответа
            min_price = extract_min_price_from_response(xml_data)

            # Если минимальная цена найдена, записываем её в Excel
            if min_price is not None:
                work_sheet.cell(row=row_idx, column=price_column).value = min_price
            else:
                work_sheet.cell(row=row_idx,
                                column=price_column).value = old_price  # Если цена не найдена записываем старую цену
        else:
            work_sheet.cell(row=row_idx,
                            column=price_column).value = old_price  # Если данных нет, записываем старую цену

        # Интервал между запросами
        time.sleep(interval)

    # Сохраняем изменения в тот же файл
    try:
        workbook.save(input_file)  # Сохраняем в исходный файл
        print(f'Результаты сохранены в {input_file}')
    except Exception as e:
        print(f'Ошибка записи файла: {e}')


# Точка входа
if __name__ == '__main__':
    # Укажите путь к файлу с артикулами и брендами
    input_file = 'data/data.xlsx'  # Входной файл
    api_key = 'MBmE7rdJlQjqwrJABOMhMqxuTZvHFnXB9wRqhhEq9QfhXGaeiFB1bN0nyzl'
    interval = 3  # Интервал между запросами в секундах

    # region = get_regions(api_key=api_key)

    process_excel(input_file=input_file, api_key=api_key, interval=interval)
