import time
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

from data.data import login, password


# Функция для отправки одного запроса
def get_data_from_api(login: str, password: str, code: str, brand: str, force_online: str = "0",
                      crosses: str = "disallow", force_diler_replace: str = "0") -> str | None:
    # URL API
    url = "https://xml.adeo.pro/pricedetals2.php"

    # Формируем XML-запрос
    xml = f'''<?xml version="1.0" encoding="UTF-8"?>
    <message>
      <param>
        <action>price</action>
        <login>{login}</login>
        <password>{password}</password>
        <code>{code}</code>
        <brand>{brand}</brand>
        <crosses>{crosses}</crosses>
        <force_online>{force_online}</force_online>
        <force_dilerReplace>{force_diler_replace}</force_dilerReplace>
      </param>
    </message>'''

    try:
        # Отправляем запрос к API
        response = requests.post(url=url, data={'xml': xml}, timeout=60)

        # Проверяем успешность запроса
        if response.status_code == 200:
            return response.text  # Возвращаем текст ответа
        else:
            print(f'Ошибка {response.status_code}: {response.text}')
            return None
    except requests.exceptions.RequestException as ex:
        print(f'Ошибка подключения: {ex}')
        return None


# Функция для извлечения минимальной цены из XML ответа, пропуская поставщиков с "Cella2108" в stock
def extract_min_price_from_response(xml_data):
    try:
        root = ET.fromstring(xml_data)
        prices = []

        for detail in root.findall('detail'):
            stock_element = detail.find('stock')
            price_element = detail.find('price')

            # Проверяем, есть ли `stock`, и пропускаем записи, содержащие "Cella2108"
            if stock_element is not None and "Cella2108" in stock_element.text:
                continue

            # Извлекаем цену, если она существует и является числом
            if price_element is not None:
                try:
                    prices.append(float(price_element.text))
                except ValueError:
                    continue  # Если значение цены некорректное, пропускаем эту запись

        return min(prices) if prices else None

    except Exception as ex:
        print(f'Ошибка при парсинге XML: {ex}')
        return None


# Основная функция
def process_excel(input_file: str, interval: int = 5, value_percent: int = 80):
    try:
        workbook = load_workbook(input_file)
        work_sheet = workbook.active  # Доступ к активному листу
    except Exception as ex:
        print(f'Ошибка чтения файла: {ex}')
        return

    # Получаем заголовки из 10-й строки
    headers = [cell.value for cell in work_sheet[11]]  # Заголовки столбцов находятся в 11-й строке

    # Индексы столбцов для "Номенклатура.Производитель", "Артикул", "Цена" по названиям
    try:
        # brand_column = headers.index('Номенклатура.Производитель') + 1  # Индексы в openpyxl начинаются с 1
        # code_column = headers.index('Артикул') + 1
        # old_price_column = headers.index('Цена') + 1
        brand_column = 1
        code_column = 4
        old_price_column = 10
        price_column = 11
    except ValueError as e:
        print(f'Ошибка: не найдены необходимые столбцы. {e}')
        return

    # Обрабатываем строки (начиная с 12-й строки, т.к. 11-я - это заголовки)
    for row_idx in range(12, work_sheet.max_row + 1):
        brand = work_sheet.cell(row=row_idx, column=brand_column).value
        code = work_sheet.cell(row=row_idx, column=code_column).value
        old_price = work_sheet.cell(row=row_idx, column=old_price_column).value

        if brand is None or code is None:
            continue

        # Выполняем запрос к API
        print(f'Отправляем запрос для артикула: {code}, бренд: {brand}')
        xml_data = get_data_from_api(login=login, password=password, code=code, brand=brand)

        if xml_data:
            # Извлекаем минимальную цену из ответа
            min_price = extract_min_price_from_response(xml_data)

            # Если минимальная цена найдена, записываем её в Excel
            if min_price is not None:
                min_price = round(min_price * (value_percent / 100), 2)
                work_sheet.cell(row=row_idx, column=price_column).value = min_price
            else:
                work_sheet.cell(row=row_idx,
                                column=price_column).value = old_price  # Если цена не найдена записываем старую цену
        else:
            work_sheet.cell(row=row_idx,
                            column=price_column).value = old_price  # Если данных нет, записываем старую цену

        # Интервал между запросами
        time.sleep(interval)

    # Сохраняем изменения в тот же файл
    try:
        workbook.save(input_file)  # Сохраняем в исходный файл
        print(f'Результаты сохранены в {input_file}')
    except Exception as ex:
        print(f'Ошибка записи файла: {ex}')


if __name__ == '__main__':
    try:
        user_input = input('Введите значение процентов от цены (по умолчанию 80): ')
        if user_input == '':
            value_percent = 80
        else:
            value_percent = int(user_input)
    except ValueError:
        raise ValueError(
            '''Введено неправильное значение. Значение должно быть целым числом или 
            нажмите Enter для установки значения по умолчанию.''')

    # Укажите путь к файлу с артикулами и брендами
    input_file = 'data/data.xlsx'  # Входной файл
    interval = 2  # Интервал между запросами в секундах

    process_excel(input_file=input_file, interval=interval, value_percent=value_percent)
