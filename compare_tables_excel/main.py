import os
import glob
from datetime import datetime

import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Цвета для выделения изменений
ADDED_FILL = PatternFill(start_color="8ED98E", end_color="8ED98E", fill_type="solid")   # Зелёный — добавленные строки
REMOVED_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")   # Красный — удалённые строки
CHANGED_FILL = PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid")   # Жёлтый — изменённые ячейки


# Цвета для процентных изменений цены
PRICE_CHANGE_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # 1-3%
PRICE_CHANGE_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 3-5%
PRICE_CHANGE_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # >5%

price_columns = [
    "Цена при заказе от 10 000 руб.",
    "Цена при заказе от 100 000 руб.",
    "Цена при заказе от 300 000 руб."
]

def load_excel(path: str) -> DataFrame:
    df = pd.read_excel(path, dtype=str)
    df.columns = df.columns.str.strip()
    return df.fillna('')


def compare_data(old_df: DataFrame, new_df: DataFrame, key_column: str = 'Артикул') -> tuple:
    old_df = old_df.set_index(key_column)
    new_df = new_df.set_index(key_column)

    added_keys = new_df.index.difference(old_df.index)
    removed_keys = old_df.index.difference(new_df.index)
    common_keys = new_df.index.intersection(old_df.index)

    changes: dict[str, dict[str, bool]] = {}

    for key in common_keys:
        row_changes = {}
        for col in new_df.columns:
            old_val = old_df.at[key, col] if col in old_df.columns else ''
            new_val = new_df.at[key, col] if col in new_df.columns else ''
            if str(old_val).strip() != str(new_val).strip():
                row_changes[col] = True
        if row_changes:
            changes[key] = row_changes

    return added_keys, removed_keys, changes


def add_price_change_columns_multiple(old_df: DataFrame, new_df: DataFrame, price_columns: list, key_column: str = 'Артикул') -> DataFrame:
    old_df = old_df.set_index(key_column)
    new_df = new_df.set_index(key_column)

    for price_col in price_columns:
        # Имя колонок для первоначальной цены и % изменения
        suffix = price_col.split("от")[-1].strip()  # например "10 000 руб."
        base_col = f'Первоначальная цена ({suffix})'
        change_col = f'Изменение цены % ({suffix})'

        new_df[base_col] = ''
        new_df[change_col] = ''

        for key in new_df.index:
            if key in old_df.index:
                old_price_str = old_df.at[key, price_col] if price_col in old_df.columns else ''
                new_price_str = new_df.at[key, price_col] if price_col in new_df.columns else ''
                try:
                    old_price = float(str(old_price_str).replace(',', '.'))
                    new_price = float(str(new_price_str).replace(',', '.'))
                    new_df.at[key, base_col] = old_price
                    if old_price != 0:
                        percent_change = (new_price - old_price) / old_price * 100
                        new_df.at[key, change_col] = round(percent_change, 2)
                    else:
                        new_df.at[key, change_col] = 0
                except Exception:
                    new_df.at[key, base_col] = ''
                    new_df.at[key, change_col] = ''
            else:
                new_df.at[key, base_col] = ''
                new_df.at[key, change_col] = ''

    new_df = new_df.reset_index()
    return new_df


def apply_formatting(result_path: str, added: list, removed: list, changed: dict, key_column: str = 'Артикул') -> None:
    wb = load_workbook(result_path)
    ws = wb.active

    header = {cell.value: idx for idx, cell in enumerate(ws[1])}
    key_idx = header.get(key_column)

    # Закрашиваем добавленные и изменённые строки
    for row in ws.iter_rows(min_row=2):
        key = str(row[key_idx].value).strip()
        if key in added:
            for cell in row:
                cell.fill = ADDED_FILL
        elif key in changed:
            for col, is_changed in changed[key].items():
                if is_changed:
                    col_idx = header.get(col)
                    if col_idx is not None:
                        row[col_idx].fill = CHANGED_FILL

        # Проверяем колонки с процентом изменения цены и выделяем цветом по диапазонам
        for price_col in price_columns:
            suffix = price_col.split("от")[-1].strip()
            change_col_name = f'Изменение цены % ({suffix})'
            col_idx = header.get(change_col_name)
            if col_idx is not None:
                cell = row[col_idx]
                try:
                    val = float(cell.value)
                    abs_val = abs(val)
                    if 1 <= abs_val < 3:
                        cell.fill = PRICE_CHANGE_GREEN
                    elif 3 <= abs_val < 5:
                        cell.fill = PRICE_CHANGE_YELLOW
                    elif abs_val >= 5:
                        cell.fill = PRICE_CHANGE_RED
                except Exception:
                    pass

    # Добавляем удалённые строки в конец таблицы
    last_row = ws.max_row + 1
    for key in removed:
        ws.cell(row=last_row, column=key_idx + 1, value=key).fill = REMOVED_FILL
        for col, idx in header.items():
            if col != key_column:
                ws.cell(row=last_row, column=idx + 1, value='(Удалён)').fill = REMOVED_FILL
        last_row += 1

    wb.save(result_path)


def save_result(new_df: DataFrame, output_dir: str = 'results') -> str:
    now_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"compare_tables_{now_str}.xlsx"
    output_path = os.path.join(output_dir, filename)

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    new_df.to_excel(output_path, index=False)
    return output_path


def main() -> None:
    folder = 'data'
    excel_files = sorted(glob.glob(os.path.join(folder, '*.xlsx')))
    if len(excel_files) < 2:
        print("❌ Не найдено минимум два .xlsx файла в папке 'data'")
        return

    old_path = excel_files[0]
    new_path = excel_files[1]

    print('📁 Загружаю таблицы...')
    old_df = load_excel(old_path)
    new_df = load_excel(new_path)

    print('🔍 Сравнение данных...')
    added, removed, changed = compare_data(old_df, new_df)

    print('➕ Вычисляю изменения цен...')
    new_df = add_price_change_columns_multiple(old_df, new_df, price_columns=price_columns)

    print('💾 Сохраняю результат...')
    result_file = save_result(new_df)

    print('🎨 Применяю форматирование...')
    apply_formatting(result_file, added, removed, changed)

    print(f'✅ Готово! Результат: {result_file}')


if __name__ == '__main__':
    main()
