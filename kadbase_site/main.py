# ===============================
# 📄 Импорты
# ===============================
import os
import glob
import time
from datetime import datetime
import xml.etree.ElementTree as ET

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

from pandas import DataFrame, ExcelWriter
from openpyxl import load_workbook


# ===============================
# 🧭 Инициализация браузера
# ===============================
def init_undetected_chromedriver(headless_mode: bool = False):
    """
    Инициализирует браузер Chrome с использованием undetected_chromedriver.

    :param headless_mode: если True — браузер запустится без интерфейса.
    :return: объект WebDriver.
    """
    options = ChromeOptions()
    if headless_mode:
        options.add_argument('--headless')

    driver = uc.Chrome(options=options)
    driver.implicitly_wait(3)
    driver.maximize_window()
    return driver


# ===============================
# 📂 Работа с файлами XML
# ===============================
def get_xml_files(folder: str) -> list[str]:
    """
    Возвращает список всех XML-файлов в указанной папке.
    """
    return glob.glob(os.path.join(folder, '*.xml*'))


def cad_number_to_tuple(cad_number: str | None) -> tuple[int, ...] | None:
    """
    Преобразует кадастровый номер в кортеж чисел для корректного сравнения.
    """
    if not cad_number:
        return None

    try:
        return tuple(int(part) for part in cad_number.split(':'))
    except ValueError:
        return None


def extract_land_records(root: ET.Element) -> list[dict[str, str | None]]:
    """
    Извлекает данные о земельных участках из XML-дерева.
    """
    records: list[dict[str, str | None]] = []
    min_cad_number = cad_number_to_tuple('38:27:000152:460')

    for land_record in root.findall('.//land_record'):
        obj = land_record.find('./object')
        if obj is None:
            continue

        # Проверяем, что объект — земельный участок
        type_el = obj.find('./common_data/type/value')
        if type_el is None or type_el.text != 'Земельный участок':
            continue

        # Извлекаем необходимые поля
        cad_number = (obj.findtext('./common_data/cad_number') or '').strip() or None
        current_cad_number = cad_number_to_tuple(cad_number)
        if current_cad_number is None or min_cad_number is None or current_cad_number < min_cad_number:
            continue

        address = (land_record.findtext('./address_location/address/readable_address') or '').strip() or None
        area_val = land_record.findtext('./params/area/value')
        area_unit = land_record.findtext('./params/area/unit')

        # Если основной площади нет — пробуем получить из другого блока
        if not area_val:
            area_val = land_record.findtext('../../area_quarter/area')
            area_unit = land_record.findtext('../../area_quarter/unit')

        area = f"{area_val.strip()} {area_unit.strip()}" if area_val and area_unit else (
            area_val.strip() if area_val else None
        )

        category = (land_record.findtext('./params/category/type/value') or '').strip() or None
        permitted_use = (land_record.findtext(
            './params/permitted_use/permitted_use_established/by_document') or '').strip() or None
        cad_cost = (land_record.findtext('./cost/value') or '').strip() or None

        records.append({
            'Кадастровый номер': cad_number,
            'Вид объекта недвижимости': 'Земельный участок',
            'Адрес': address,
            'Площадь или основная характеристика': area,
            'Категория земель': category,
            'Вид разрешенного использования': permitted_use,
            'Кадастровая стоимость': cad_cost,
            'Форма собственности': ''
        })
    return records


def get_result_file_name(file_path: str) -> str:
    """
    Возвращает имя Excel-файла на основе имени входного XML-файла.
    """
    file_name = os.path.basename(file_path)
    xml_index = file_name.lower().find('.xml')
    if xml_index != -1:
        return f'{file_name[:xml_index]}.xlsx'
    return f'{os.path.splitext(file_name)[0]}.xlsx'


def parse_xml_file(folder: str) -> None:
    """
    Обрабатывает все XML-файлы в указанной папке.
    """
    xml_files: list[str] = get_xml_files(folder)

    for file_path in xml_files:
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            records = extract_land_records(root)
            save_excel(records, get_result_file_name(file_path))
        except ET.ParseError as e:
            print(f'❌ Ошибка парсинга файла {file_path}: {e}')

        print(f'Обработан файл: {file_path} ')


# ===============================
# 📊 Сохранение в Excel
# ===============================
def save_excel(data: list[dict[str, str | None]], file_name: str) -> None:
    """
    Сохраняет данные в Excel-файл в папке 'results'.
    """
    directory = 'results'
    os.makedirs(directory, exist_ok=True)
    file_path: str = os.path.join(directory, file_name)

    df = DataFrame(data)
    with ExcelWriter(file_path, mode='w') as writer:
        df.to_excel(writer, sheet_name='Лист1', index=False)

    print(f'✅ Данные сохранены в файл: {file_path}')


# ===============================
# 🏢 Обновление столбца "Форма собственности"
# ===============================
def update_ownership_excel(driver, excel_path: str, url: str, sheet: str = 'Лист2', batch_size: int = 100) -> None:
    """
    Обновляет данные в Excel: проверяет форму собственности по кадастровому номеру
    через сайт и удаляет строки с частной формой собственности или при отсутствии данных.

    Данные сохраняются каждые batch_size обработанных строк.
    """
    wb = load_workbook(excel_path)
    ws = wb[sheet]

    headers: list[str] = [cell.value for cell in ws[1]]
    cad_col: int = headers.index('Кадастровый номер') + 1
    ownership_col: int = headers.index('Форма собственности') + 1

    cad_rows: list[int] = [row for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=cad_col).value]
    total: int = len(cad_rows)
    processed_count: int = 0

    print(f"🔍 Всего строк для проверки: {total}")

    # Обходим строки снизу вверх (чтобы безопасно удалять)
    for row in reversed(cad_rows):
        cad_number: str = str(ws.cell(row=row, column=cad_col).value)
        print(f'➡️  Обработка {processed_count}/{total} — {cad_number}')

        try:
            driver.get(url)

            # Ожидаем поле ввода
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input.input_search"))
            )

            # Вводим кадастровый номер
            input_box = driver.find_element(By.CSS_SELECTOR, 'input.input_search')
            input_box.clear()
            input_box.send_keys(cad_number)
            input_box.send_keys(Keys.ENTER)

            # Проверяем наличие блока с формой собственности
            try:
                ownership_div = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//div[contains(text(),'Форма собственности')]")
                    )
                )

                # Извлекаем текст из соседнего блока
                try:
                    ownership = ownership_div.find_element(By.XPATH, "following-sibling::div").text.strip()
                except Exception:
                    ws.delete_rows(row)
                    continue
            except TimeoutException:
                ws.delete_rows(row)
                continue

            # Удаляем строку, если собственность частная
            ownership = ownership.strip().lower()
            if ownership == 'частная':
                ws.delete_rows(row)
            else:
                processed_count += 1
                ws.cell(row=row, column=ownership_col, value=ownership)

        except Exception as ex:
            print(f'❌ Ошибка при обработке {cad_number}: {ex}')

        # Промежуточное сохранение каждые batch_size строк
        if processed_count % batch_size == 0:
            wb.save(excel_path)
            print(f'💾 Промежуточное сохранение: обработано {processed_count}/{total}')

    # Финальное сохранение
    wb.save(excel_path)
    print(f'✅ Excel обновлён: {excel_path}')


def update_restrictions_excel(driver, excel_path: str, url: str, sheet: str = 'Лист1', batch_size: int = 10) -> None:
    """
    Обновляет Excel, оставляя только строки, где на сайте по кадастровому номеру
    найден текст:
    "Зарегистрированные ограничения - не обнаружены (на момент последней проверки)"

    Все остальные строки удаляются.
    """

    wb = load_workbook(excel_path)
    ws = wb[sheet]

    headers: list[str] = [cell.value for cell in ws[1]]
    cad_col: int = headers.index('Кадастровый номер') + 1

    cad_rows: list[int] = [row for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=cad_col).value]
    total: int = len(cad_rows)
    processed_count: int = 0

    print(f"🔍 Всего строк для проверки: {total}")

    for row in reversed(cad_rows):
        cad_number: str = str(ws.cell(row=row, column=cad_col).value)
        print(f'➡️  Обработка {processed_count}/{total} — {cad_number}')

        try:
            driver.get(url)

            # Ожидаем появления поля ввода
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input.input_search"))
            )

            # Вводим кадастровый номер
            input_box = driver.find_element(By.CSS_SELECTOR, 'input.input_search')
            input_box.clear()
            input_box.send_keys(cad_number)
            input_box.send_keys(Keys.ENTER)

            # Проверяем наличие нужного блока
            try:
                restriction_div = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located(
                        (By.XPATH,
                         "//div[contains(@class, 'left_sser') and contains(text(), 'Зарегистрированные ограничения - не обнаружены')]"
                         )
                    )
                )
                restriction_text = restriction_div.text.strip()
            except TimeoutException:
                ws.delete_rows(row)
                continue

            # Проверяем, совпадает ли текст
            if "Зарегистрированные ограничения - не обнаружены" in restriction_text:
                processed_count += 1
            else:
                ws.delete_rows(row)

        except Exception as ex:
            print(f'❌ Ошибка при обработке {cad_number}: {ex}')
            ws.delete_rows(row)

        # Промежуточное сохранение каждые batch_size строк
        if processed_count % batch_size == 0:
            wb.save(excel_path)
            print(f'💾 Промежуточное сохранение: обработано {processed_count}/{total}')

    wb.save(excel_path)
    print(f'✅ Excel обновлён: {excel_path}')


# ===============================
# 🚀 Точка входа
# ===============================
def main() -> None:
    """
    Основная точка входа в программу:

    - Формирует текущую дату `cur_date` для имени выходного Excel-файла.
    - Загружает XML-данные из папки `data`.
    - Сохраняет данные в Excel с именем, включающим `cur_date`.
    - Инициализирует браузер Chrome через undetected_chromedriver.
    - Переходит на сайт и предоставляет пользователю 60 секунд на авторизацию.
    - Обновляет Excel-файл: проверяет форму собственности по кадастровым номерам
      и удаляет строки с частной формой собственности или при отсутствии данных.
    - В конце выводит общее время выполнения программы.
    """
    start_time = datetime.now()

    cur_date = datetime.now().strftime('%d-%m-%Y')

    data_folder: str = 'data'
    url: str = 'https://kadbase.ru/'

    # Если нужно — распарсить XML и создать Excel
    parse_xml_file(data_folder)
    # excel_path: str = f'results/result_data_{cur_date}.xlsx'
    #
    # driver = init_undetected_chromedriver(headless_mode=False)
    # try:
    #     # Авторизация вручную
    #     driver.get("https://kadbase.ru/lk/")
    #     time.sleep(60)
    #     print("⏳ У вас есть 60 секунд, чтобы авторизоваться вручную...")
    #
    #     # update_ownership_excel(driver, excel_path, url)
    #     update_restrictions_excel(driver, excel_path, url)
    # finally:
    #     driver.close()
    #     driver.quit()

    execution_time = datetime.now() - start_time
    print('Сбор данных завершен!')
    print(f'Время работы программы: {execution_time}')


if __name__ == '__main__':
    main()
