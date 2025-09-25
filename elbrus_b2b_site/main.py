import os
import time
import glob
import re
from datetime import datetime

from requests import Session
from pandas import DataFrame, ExcelWriter
from openpyxl import load_workbook

from bs4 import BeautifulSoup

start_time: datetime = datetime.now()


def get_html(url: str, headers: dict, session: Session) -> str | None:
    """
    Выполняет HTTP GET-запрос и возвращает HTML-код страницы.
    """
    try:
        response: 'requests.Response' = session.get(url=url, headers=headers, timeout=60)

        if response.status_code != 200:
            print(f'status_code: {response.status_code}')

        html: str = response.text
        return html
    except Exception as ex:
        print(f'get_html: {ex}')
        return None


def get_normalize_article(article) -> str | None:
    """
    Приводит артикул к единому виду.
    """
    if article is None:
        return None

    article: str = str(article)

    if '•' in article:
        article = article.split('•')[-1]
    if article.endswith('-L'):
        article = article[:-2]

    article = re.sub(r'[\s\-/]', '', article)
    return article.strip() or None


def extract_value(soup: BeautifulSoup, label: str) -> float | None:
    """
    Извлекает числовое значение упаковки из HTML в метрах.
    """
    try:
        value_text: str = soup.find('span', string=re.compile(label)).find_next().get_text(strip=True)
        return float(value_text.replace(',', '.')) / 1000
    except Exception:
        return None


def save_excel(data: list[dict], species: str) -> None:
    """
    Сохраняет список словарей в Excel-файл.
    """
    directory: str = 'results'
    os.makedirs(directory, exist_ok=True)

    file_path: str = f'{directory}/result_data_{species}.xlsx'
    dataframe: DataFrame = DataFrame(data)

    with ExcelWriter(file_path, mode='w') as writer:
        dataframe.to_excel(writer, sheet_name='data', index=False)

    print(f'Данные сохранены в файл {file_path}')


def get_product_data(article: str, headers: dict, session: Session) -> float | None:
    """
    Извлекает объём упаковки конкретного артикула с сайта.
    """
    product_url: str = f"https://elbrus-b2b.ru/search?pcode={article}"

    try:
        time.sleep(1)
        html: str | None = get_html(url=product_url, headers=headers, session=session)
    except Exception as ex:
        html = None
        print(f'get_products_urls: {product_url} - {ex}')

    if not html:
        print(f'not html product_url: {product_url}')
        return None

    soup: BeautifulSoup = BeautifulSoup(html, 'lxml')

    try:
        short_url: str = soup.find('div', class_='searchTopLinks').find_next_sibling('a').get('href')
        review_url: str = f'https://elbrus-b2b.ru{short_url}'
    except Exception as ex:
        if soup.find('img', class_='captchaImg'):
            print(f'[CAPTCHA] Обнаружена капча на {product_url}')
        else:
            print(f'not short_url: {product_url} - {ex}')
        return None

    try:
        time.sleep(1)
        html = get_html(url=review_url, headers=headers, session=session)
    except Exception as ex:
        html = None
        print(f'get_products_urls: {product_url} - {ex}')

    if not html:
        print(f'not html review_url: {product_url}')
        return None

    soup = BeautifulSoup(html, 'lxml')

    height_m: float | None = extract_value(soup, 'Высота упаковки, мм:')
    length_m: float | None = extract_value(soup, 'Длина упаковки, мм:')
    width_m: float | None = extract_value(soup, 'Ширина упаковки, мм:')

    if None in (height_m, length_m, width_m):
        print(f'not volume: {product_url}')
        return None

    volume_m3: float = height_m * length_m * width_m
    return volume_m3


def process_data_files(data_folder: str = 'data') -> None:
    """
    Обрабатывает все Excel-файлы в указанной папке.
    """
    excel_files: list[str] = glob.glob(os.path.join(data_folder, "*.xls*"))

    headers: dict = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'priority': 'u=0, i',
        'referer': 'https://elbrus-b2b.ru/search/O.E.M./OEM0003KBPL',
        'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
    }

    with Session() as session:
        for data_file_path in excel_files:
            file_name: str = os.path.basename(data_file_path)
            print(f'[INFO] Обрабатывается файл: {file_name}')

            try:
                wb = load_workbook(data_file_path)
            except Exception as e:
                print(f'[ERROR] Ошибка чтения файла {file_name}: {e}')
                continue

            for sheet_name in wb.sheetnames[9:10]:
                ws = wb[sheet_name]
                headers_excel: list[str] = [cell.value for cell in ws[4]]

                volume_column_index: int = headers_excel.index('Volume')

                for row in ws.iter_rows(min_row=5, max_col=len(headers_excel)):
                    article_cell = row[1].value
                    if not article_cell:
                        continue
                    article: str | None = get_normalize_article(article_cell)

                    volume_m3: float | None = get_product_data(article=article, headers=headers, session=session)

                    if not volume_m3:
                        continue

                    row[volume_column_index].value = volume_m3
                    print(f'Обработан артикул: {article}')

                wb.save(data_file_path)
                print(f"[INFO] Лист '{sheet_name}' сохранён")


def main() -> None:
    """
    Основная функция запуска скрипта.
    """
    process_data_files()
    execution_time = datetime.now() - start_time
    print('Сбор данных завершен.')
    print(f'Время выполнения: {execution_time}')


if __name__ == '__main__':
    main()
