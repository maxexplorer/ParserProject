import os
import time
import math
import glob
import re
from datetime import datetime
from urllib.parse import urlparse, parse_qs

from requests import Session
from pandas import DataFrame, ExcelWriter, read_excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

start_time = datetime.now()


def get_seller_id(url: str) -> str | None:
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    seller_id = query_params.get("sellerId")
    return seller_id[0] if seller_id else None


def get_normalize_article(article) -> str | None:
    if article is None:
        return None

    article = str(article)  # всегда строка

    if '•' in article:
        article = article.split('•')[-1]
    if article.endswith('-L'):
        article = article[:-2]

    # убираем пробелы, дефисы и слэши
    article = re.sub(r'[\s\-/]', '', article)

    return article.strip() or None


def read_excel_data(file_path: str, own: bool = False) -> dict:
    """
    Чтение данных из Excel файла Avito.
    own=True → возвращает {артикул: цена}
    own=False → возвращает {артикул: {'Название', 'Бренд', 'Цена', 'Ссылка'}}
    """
    df = read_excel(file_path, header=0, sheet_name=0)  # считаем заголовки из первой строки
    data_dict = {}

    if own:
        # Ожидаем колонки: ['Артикул', 'Наша цена']
        for _, row in df.iterrows():
            article = str(row[df.columns[0]]).replace('-', '').strip()
            price = row[df.columns[1]]
            if article:
                data_dict[article] = price
    else:
        # Ожидаем колонки: ['Артикул', 'Название', 'Бренд', 'Цена', 'Ссылка']
        for _, row in df.iterrows():
            article = str(row[df.columns[1]]).replace('-', '').strip()
            title = row.get('Название')
            brand = row.get('Бренд')
            price = row.get('Цена')
            url = row.get('Ссылка')
            if article:
                data_dict[article] = {
                    "Название": title,
                    "Бренд": brand,
                    "Цена": price,
                    "Ссылка": url
                }

    return data_dict


# Функция для записи данных в формат xlsx
def save_excel(data: list, species: str) -> None:
    directory = 'results'

    os.makedirs(directory, exist_ok=True)

    file_path = f'{directory}/result_data_{species}.xlsx'

    dataframe = DataFrame(data)

    with ExcelWriter(file_path, mode='w') as writer:
        dataframe.to_excel(writer, sheet_name='data', index=False)

    print(f'Данные сохранены в файл {file_path}')


def get_products_data(seller_url: str, limit: int = 100, own: bool = False, species: str = None):
    """
    Сбор данных по Avito.
    Если own=True → сохраняем {артикул: цена} в own_result.xlsx
    Если own=False → сохраняем {артикул: {...}} в competitor_data.xlsx
    """
    seller_id = get_seller_id(seller_url)
    if not seller_id:
        raise ValueError("Не удалось извлечь sellerId")

    with Session() as session:
        result_list = []

        headers = {
            'accept': 'application/json',
            'referer': seller_url,
            'user-agent': 'Mozilla/5.0',
            'x-requested-with': 'XMLHttpRequest',
        }

        first_params = {
            'p': 1,
            'sellerId': seller_id,
            'itemsOnPage': limit,
            'limit': limit}

        response = session.get('https://www.avito.ru/web/1/profile/items',
                               params=first_params, headers=headers, timeout=(3, 5))
        response.raise_for_status()
        json_data = response.json()

        total = json_data.get('foundCount', 0)
        if total == 0:
            print(f"{seller_url}: товаров нет")
            return

        pages = math.ceil(total / limit)
        print(f"{seller_url}: всего {total} товаров, {pages} страниц")

        # цикл по страницам
        for page in range(1, pages + 1):
            params = first_params.copy()
            params['p'] = page

            try:
                time.sleep(1)
                response = session.get(
                    'https://www.avito.ru/web/1/profile/items',
                    params=params,
                    headers=headers,
                    timeout=(3, 5)
                )

                if response.status_code != 200:
                    print(f' seller_url: {seller_url}: статус ответа {response.status_code}')
                    continue

                json_data: dict = response.json()
                items: list = json_data.get('catalog', {}).get('items', [])

            except Exception as ex:
                print(f"{seller_url} страница {page}: {ex}")
                continue

            if not items:
                continue

            for item in items:
                product_url = f"https://www.avito.ru{item.get('urlPath')}"
                title = item.get('title')

                article_list = item.get('iva', {}).get('SparePartsParamsStep') or []
                article = article_list[0].get('payload', {}).get('text') if article_list else None
                normalize_article = get_normalize_article(article)
                if not normalize_article:
                    continue

                prices = item.get('priceDetailed')
                if not prices:
                    continue
                price = prices.get('value')

                if own:
                    result_list.append(
                        {
                            'Артикул': normalize_article,
                            'Цена': price,
                        }
                    )
                else:
                    brand_list = item.get('iva', {}).get('AutoPartsManufacturerStep') or []
                    brand = brand_list[0].get('payload', {}).get('value') if brand_list else None
                    result_list.append(
                        {
                            'Название': title,
                            'Артикул': normalize_article,
                            'Бренд': brand,
                            'Цена': price,
                            'Ссылка': product_url
                        }
                    )

            print(f"Обработано страниц: {page}/{pages}")

        print(f"{seller_url}: обработано {len(result_list)} записей")

    save_excel(result_list, species)


def process_data_files(data_folder='data'):
    # Загружаем ранее сохранённые Excel
    own_dict = read_excel_data('results/result_data_own.xlsx', own=True)
    competitor_dict = read_excel_data('results/result_data_competitor.xlsx', own=False)

    excel_files = glob.glob(os.path.join(data_folder, "*.xls*"))

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for data_file_path in excel_files:
        file_name = os.path.basename(data_file_path)
        print(f'[INFO] Обрабатывается файл: {file_name}')

        try:
            wb = load_workbook(data_file_path)
        except Exception as e:
            print(f'[ERROR] Ошибка чтения файла {file_name}: {e}')
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[2]]

            needed_cols = ['Бренд', 'Цена', 'Наша цена', 'Ссылка']
            col_indices = {}
            for col in needed_cols:
                if col not in headers:
                    ws.cell(row=2, column=len(headers) + 1, value=col)
                    headers.append(col)
                col_indices[col] = headers.index(col) + 1

            for row in ws.iter_rows(min_row=3, max_col=len(headers)):
                article_cell = row[1].value
                if not article_cell:
                    continue
                article = get_normalize_article(article_cell)

                own_price = own_dict.get(article)
                competitor_data = competitor_dict.get(article)

                if own_price:
                    ws.cell(row=row[0].row, column=col_indices['Наша цена'], value=own_price)

                if competitor_data:
                    ws.cell(row=row[0].row, column=col_indices['Бренд'], value=competitor_data['Бренд'])
                    ws.cell(row=row[0].row, column=col_indices['Цена'], value=competitor_data['Цена'])
                    ws.cell(row=row[0].row, column=col_indices['Ссылка'], value=competitor_data['Ссылка'])

                    if own_price and competitor_data['Цена']:
                        if own_price < competitor_data['Цена']:
                            ws.cell(row=row[0].row, column=col_indices['Цена']).fill = green_fill
                        elif own_price > competitor_data['Цена']:
                            ws.cell(row=row[0].row, column=col_indices['Цена']).fill = red_fill
                        else:
                            ws.cell(row=row[0].row, column=col_indices['Цена']).fill = yellow_fill

        wb.save(data_file_path)
        print(f"[INFO] Файл обновлён: {file_name}")


def main():
    # own_url = "https://www.avito.ru/brands/gg_auto/all/zapchasti_i_aksessuary?sellerId=de08ae472d1d705873cca3d2970af199"
    # print("[INFO] Сбор данных своего магазина...")
    # get_products_data(seller_url=own_url, own=True, species='own')

    # competitor_url = input("Введите ссылку на магазин конкурента: \n").strip()
    # print("[INFO] Сбор данных конкурента...")
    # get_products_data(seller_url=competitor_url, own=False, species='competitor')

    process_data_files()

    execution_time = datetime.now() - start_time
    print('Сбор данных завершен.')
    print(f'Время выполнения: {execution_time}')


if __name__ == '__main__':
    main()
