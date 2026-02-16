# propertyfinder_parser.py

import os
import time
from datetime import datetime, date, timedelta
import re

from requests import Session

from openpyxl import Workbook, load_workbook

# =============================================================================
# Глобальные настройки
# =============================================================================

# Время старта программы (используется для подсчёта времени выполнения)
start_time: datetime = datetime.now()

# HTTP-заголовки для имитации запроса от браузера
headers: dict = {
    'accept': '*/*',
    'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'priority': 'u=1, i',
    'referer': 'https://www.propertyfinder.ae/en/buy/properties-for-sale.html',
    'sec-ch-ua': '"Not(A:Brand";v="8", "Chromium";v="144", "Google Chrome";v="144"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
    'x-nextjs-data': '1',
}


def get_build_id(session: Session, headers: dict) -> str | None:
    url = 'https://www.propertyfinder.ae/en/buy/properties-for-sale.html'
    r = session.get(url, headers=headers, timeout=30)

    if r.status_code != 200:
        return None

    match = re.search(r'"buildId":"([^"]+)"', r.text)
    return match.group(1) if match else None


# =============================================================================
# Работа с API PropertyFinder
# =============================================================================

def get_json(session: Session, headers: dict, page: int) -> dict | None:
    build_id = get_build_id(session, headers)
    if not build_id:
        print('Не удалось получить buildId')
        return None

    url = f'https://www.propertyfinder.ae/search/_next/data/{build_id}/en/search.json'

    params = {
        'c': '1',
        'fu': '0',
        'ob': 'nd',
        'page': page,
    }

    response = session.get(url=url, headers=headers, params=params, timeout=30)

    if response.status_code != 200:
        print(f'status_code: {response.status_code}')
        return None

    return response.json()

def clean_text(text: str | None) -> str | None:
    if text is None:
        return None
    # Удаляем символы управления, оставляем только читаемые
    return re.sub(r'[\x00-\x1F\x7F]', ' ', text).strip()


# =============================================================================
# Сбор и обработка данных
# =============================================================================

def get_data(headers: dict, pages: int = 3, days: int = 1, batch_size: int = 100) -> list[dict[str, str | int | None]]:
    """
    Собирает объявления о продаже недвижимости с сайта PropertyFinder.

    :param headers: HTTP-заголовки запроса
    :param days: Количество дней назад (1 = только сегодня, 7 = неделя)
    :return: Список словарей с данными объявлений
    """
    result_data: list[dict[str, str | int | None]] = []

    # Граница по дате (UTC)
    today_utc: date = date.today()
    date_from: date = today_utc - timedelta(days=days - 1)

    max_consecutive_old = 3  # после 3 подряд старых объявлений выходим
    consecutive_old = 0  # общий счётчик по всем страницам

    # Используем одну HTTP-сессию для всех запросов
    with Session() as session:
        for page in range(1, pages + 1):
            try:
                time.sleep(1)
                json_data: dict | None = get_json(
                    session=session,
                    headers=headers,
                    page=page
                )

            except Exception as ex:
                print(f'page {page}: {ex}')
                continue

            if not json_data:
                print('not json_data')
                continue

            items: list | None = (
                json_data
                .get('pageProps', {})
                .get('searchResult', {})
                .get('listings')
            )

            if not items:
                print('not items')
                continue

            for item in items:
                properties: dict | None = item.get('property')
                if not properties:
                    continue

                listed_date: str | None = properties.get('listed_date')
                if not listed_date:
                    continue

                date_obj: datetime = datetime.strptime(listed_date, '%Y-%m-%dT%H:%M:%SZ')
                listed_date_only: date = date_obj.date()

                # Проверка даты
                if not (date_from <= listed_date_only <= today_utc):
                    consecutive_old += 1
                    if consecutive_old >= max_consecutive_old:
                        return result_data  # выходим из get_data полностью
                    continue
                else:
                    consecutive_old = 0  # сброс счётчика, если дата подходит

                # -----------------------------------------------------------------
                # Основные данные объекта
                # -----------------------------------------------------------------
                property_id: int | None = properties.get('id')
                property_type: str | None = properties.get('property_type')
                title: str | None = clean_text(properties.get('title'))

                # Локация
                location: dict = properties.get('location', {})
                full_name: str | None = location.get('full_name')
                building_type: str | None = location.get('type')
                building_name: str | None = location.get('name')

                # Цена
                price_info: dict = properties.get('price', {})
                price: int | None = price_info.get('value')
                currency: str | None = price_info.get('currency')

                # Комнаты
                bedrooms: int | None = properties.get('bedrooms')
                bathrooms: int | None = properties.get('bathrooms')

                # Площадь
                size_info: dict = properties.get('size', {})
                size: int | None = size_info.get('value')
                unit: str | None = size_info.get('unit')

                # Дополнительная информация
                completion_status: str | None = properties.get('completion_status')
                description: str | None = clean_text(properties.get('description'))
                amenities: str = ', '.join(clean_text(a) for a in properties.get('amenity_names', []))
                property_url: str | None = properties.get('share_url')

                # Изображения
                image_urls: str = ', '.join(
                    image.get('medium') for image in properties.get('images', []) if image.get('medium')
                )

                # Брокер
                broker: dict = properties.get('broker') or {}
                broker_name: str | None = broker.get('name')
                broker_address: str | None = broker.get('address')
                broker_email: str | None = broker.get('email')
                broker_phone: str | None = broker.get('phone')

                # -----------------------------------------------------------------
                # Добавление результата
                # -----------------------------------------------------------------
                result_data.append(
                    {
                        'listed_date': date_obj,
                        'property_id': property_id,
                        'property_type': property_type,
                        'title': title,
                        'full_name': full_name,
                        'building_type': building_type,
                        'building_name': building_name,
                        'price': price,
                        'currency': currency,
                        'bedrooms': bedrooms,
                        'bathrooms': bathrooms,
                        'size': size,
                        'size_unit': unit,
                        'completion_status': completion_status,
                        'description': description,
                        'amenities': amenities,
                        'property_url': property_url,
                        'image_urls': image_urls,
                        'broker_name': broker_name,
                        'broker_address': broker_address,
                        'broker_email': broker_email,
                        'broker_phone': broker_phone,
                    }
                )

            print(f'Обработано страниц: {page}/{pages}')

            # Сохраняем партию данных в Excel
            if len(result_data) >= batch_size:
                save_excel(result_data, today_utc)
                result_data.clear()

        # Сохраняем остаток
        if result_data:
            save_excel(result_data, today_utc)

    return result_data


# =============================================================================
# Сохранение данных
# =============================================================================
def save_excel(data: list[dict], today_utc: date) -> None:
    """
     Сохраняет список данных в Excel-файл (results/result_data_date.xlsx)
    :param data: list[dict]
    :param today_utc: date
    :return: None
    """
    cur_date: str = today_utc.strftime('%d-%m-%Y')
    directory = 'results'
    file_path = f'{directory}/result_data_{cur_date}.xlsx'

    os.makedirs(directory, exist_ok=True)

    if not os.path.exists(file_path):
        # Создаём новый файл
        wb = Workbook()
        ws = wb.active
        ws.title = 'Data'

        # Записываем заголовки
        headers = list(data[0].keys())
        ws.append(headers)

        for row in data:
            ws.append(list(row.values()))

        wb.save(file_path)
    else:
        # Открываем и дописываем строки
        wb = load_workbook(file_path)
        ws = wb['Data']

        for row in data:
            ws.append(list(row.values()))

        wb.save(file_path)

    print(f"Сохранено {len(data)} записей в {file_path}")


# =============================================================================
# Точка входа для локального тестирования
# =============================================================================

def main() -> None:
    """
    Точка входа в программу.

    Запускает сбор данных,
    сохраняет результат в Excel
    и выводит общее время выполнения.
    """

    pages: int = 800
    days_to_collect: int = 1  # 1 = сегодня, 7 = неделя
    batch_size = 100

    try:
        result_data: list = get_data(headers=headers, pages=pages, days=days_to_collect, batch_size=batch_size)

    except Exception as ex:
        print(f'main: {ex}')
        input('Нажмите Enter, чтобы закрыть программу...')

    execution_time = datetime.now() - start_time
    print('Сбор данных завершён!')
    print(f'Время работы программы: {execution_time}')


if __name__ == '__main__':
    main()
