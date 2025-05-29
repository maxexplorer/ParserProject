import os
import glob
from datetime import datetime

import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Цвета для выделения изменений
ADDED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зелёный — добавленные строки
REMOVED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Красный — удалённые строки
CHANGED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Жёлтый — изменённые ячейки


def load_excel(path: str) -> DataFrame:
    """
    Загружает Excel-файл в DataFrame, очищая пробелы в заголовках и заменяя NaN на пустые строки.

    :param path: Путь к Excel-файлу
    :return: DataFrame с очищенными данными
    """
    df = pd.read_excel(path, dtype=str)
    df.columns = df.columns.str.strip()
    return df.fillna('')


def compare_data(old_df: DataFrame, new_df: DataFrame, key_column: str = 'Артикул') -> tuple:
    """
    Сравнивает два DataFrame по ключевому столбцу, определяя добавленные, удалённые и изменённые строки.

    :param old_df: Старый DataFrame
    :param new_df: Новый DataFrame
    :param key_column: Название ключевого столбца (по умолчанию 'Артикул')
    :return: Кортеж из (добавленные ключи, удалённые ключи, словарь изменений)
    """
    old_df = old_df.set_index(key_column)
    new_df = new_df.set_index(key_column)

    added_keys = new_df.index.difference(old_df.index)
    removed_keys = old_df.index.difference(new_df.index)
    common_keys = new_df.index.intersection(old_df.index)

    changes: dict[str, dict[str, bool]] = {}

    for key in common_keys:
        row_changes = {}
        for col in new_df.columns:
            old_val = old_df.at[key, col] if col in old_df.columns else ''
            new_val = new_df.at[key, col] if col in new_df.columns else ''
            if str(old_val).strip() != str(new_val).strip():
                row_changes[col] = True
        if row_changes:
            changes[key] = row_changes

    return added_keys, removed_keys, changes


def apply_formatting(result_path: str, added: list, removed: list, changed: dict, key_column: str = 'Артикул') -> None:
    """
    Применяет цветовое форматирование к Excel-файлу на основе сравнения данных.

    :param result_path: Путь к результирующему Excel-файлу
    :param added: Список добавленных ключей
    :param removed: Список удалённых ключей
    :param changed: Словарь изменений по ключам и столбцам
    :param key_column: Название ключевого столбца (по умолчанию 'Артикул')
    """
    wb = load_workbook(result_path)
    ws = wb.active

    # Словарь соответствий названия столбца и индекса
    header = {cell.value: idx for idx, cell in enumerate(ws[1])}
    key_idx = header.get(key_column)

    # Закрашиваем добавленные и изменённые строки
    for row in ws.iter_rows(min_row=2):
        key = str(row[key_idx].value).strip()
        if key in added:
            for cell in row:
                cell.fill = ADDED_FILL
        elif key in changed:
            for col, is_changed in changed[key].items():
                if is_changed:
                    col_idx = header.get(col)
                    if col_idx is not None:
                        row[col_idx].fill = CHANGED_FILL

    # Добавляем удалённые строки в конец таблицы
    last_row = ws.max_row + 1
    for key in removed:
        ws.cell(row=last_row, column=key_idx + 1, value=key).fill = REMOVED_FILL
        for col, idx in header.items():
            if col != key_column:
                ws.cell(row=last_row, column=idx + 1, value='(Удалён)').fill = REMOVED_FILL
        last_row += 1

    wb.save(result_path)


def save_result(new_df: DataFrame, output_dir: str = 'results') -> str:
    """
    Сохраняет DataFrame в Excel-файл с названием, содержащим текущую дату и время.

    :param new_df: DataFrame для сохранения
    :param output_dir: Папка, куда сохранить результат
    :return: Путь к сохранённому файлу
    """
    now_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"result_{now_str}.xlsx"
    output_path = os.path.join(output_dir, filename)

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    new_df.to_excel(output_path, index=False)
    return output_path


def main() -> None:
    """
    Основной процесс:
    1. Загружает два Excel-файла из папки `data/`
    2. Сравнивает их по ключевому столбцу
    3. Сохраняет результат с текущей датой и временем
    4. Применяет цветовое форматирование по результатам сравнения
    """
    folder = 'data'
    excel_files = sorted(glob.glob(os.path.join(folder, '*.xlsx')))
    if len(excel_files) < 2:
        print("❌ Не найдено минимум два .xlsx файла в папке 'data'")
        return

    old_path = excel_files[0]
    new_path = excel_files[1]

    print('📁 Загружаю таблицы...')
    old_df = load_excel(old_path)
    new_df = load_excel(new_path)

    print('🔍 Сравнение данных...')
    added, removed, changed = compare_data(old_df, new_df)

    print('💾 Сохраняю результат...')
    result_file = save_result(new_df)

    print('🎨 Применяю форматирование...')
    apply_formatting(result_file, added, removed, changed)

    print(f'✅ Готово! Результат: {result_file}')


if __name__ == '__main__':
    main()
