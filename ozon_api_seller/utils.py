import os
from datetime import datetime
from io import StringIO
import glob
from copy import copy

import pandas as pd
from openpyxl import load_workbook

from data.data import COLUMN_MAPPING


def save_excel(data: list[dict], filename_prefix: str) -> None:
    now_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"{filename_prefix}_{now_str}.xlsx"
    folder = os.path.dirname(filename)
    if folder:
        os.makedirs(folder, exist_ok=True)

    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)
    print(f'✅ Данные сохранены в {filename}')


def get_target_columns_from_template(template_path: str, sheet_name: str = 'Товары и цены') -> list[str]:
    wb = load_workbook(template_path)
    ws = wb[sheet_name]
    return [cell.value for cell in ws[3] if cell.value]


def copy_cell_style(src_cell, dest_cell):
    """
    Копирует стили из src_cell в dest_cell.
    """
    dest_cell.font = copy(src_cell.font)
    dest_cell.border = copy(src_cell.border)
    dest_cell.fill = copy(src_cell.fill)
    dest_cell.number_format = copy(src_cell.number_format)
    dest_cell.alignment = copy(src_cell.alignment)


def save_to_excel_template(df: pd.DataFrame, template_path: str, output_path: str,
                           sheet_name: str = 'Товары и цены') -> None:
    """
    Записывает DataFrame в шаблон Excel, начиная с 5-й строки, копируя стили шаблонной строки.
    """
    folder = os.path.dirname(output_path)
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'❌ В шаблоне отсутствует лист "{sheet_name}"')

    ws = wb[sheet_name]

    # Шаблонная строка для копирования стилей (5-я строка в шаблоне)
    template_row_idx = 5
    template_row = [ws.cell(row=template_row_idx, column=col_idx) for col_idx in range(1, ws.max_column + 1)]

    # Записываем данные начиная с 5-й строки
    start_row = 5
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for col_idx, value in enumerate(row, start=1):
            dest_cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx <= len(template_row):
                copy_cell_style(template_row[col_idx - 1], dest_cell)

    wb.save(output_path)
    print(f'✅ Excel сохранён по шаблону: {output_path}')


def process_and_save_excel_from_csv_content(csv_content: str) -> None:
    df = pd.read_csv(StringIO(csv_content), delimiter=';')

    # Переименовываем колонки
    df.rename(columns=COLUMN_MAPPING, inplace=True)

    # Получаем нужные колонки из шаблона
    folder = 'templates'
    template_path = glob.glob(os.path.join(folder, '*.xlsx'))[0]
    target_columns = get_target_columns_from_template(template_path, sheet_name='Товары и цены')

    # Добавляем отсутствующие целевые колонки пустыми
    for col in target_columns:
        if col not in df.columns:
            df[col] = ''

    # Копируем цену
    price_col = 'Текущая цена (со скидкой), руб.'
    min_price_col = 'Минимальная цена, руб.'
    if price_col in df.columns:
        df[min_price_col] = df[price_col]

    # Убираем NaN
    df = df.fillna('')

    # Упорядочиваем колонки
    df = df[target_columns]

    # Приводим некоторые колонки к строковому типу
    str_columns = ['Артикул', 'SKU', 'Ozon SKU ID', 'Штрихкод', 'Barcode', 'Объем, л', 'Объемный вес, кг']
    for col in str_columns:
        if col in df.columns:
            df[col] = df[col].astype(str).str.lstrip("'").str.strip()

    # Сохраняем в шаблон
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_path = f'results/product_report_{timestamp}.xlsx'

    save_to_excel_template(df, template_path, output_path, sheet_name='Товары и цены')
