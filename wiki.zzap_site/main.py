import time
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

import requests


def get_regions(login: str, password: str, api_key: str):
    url = "https://api.zzap.pro/webservice/datasharing.asmx/GetRegionsV2"

    # Формируем параметры запроса
    params = {
        "login": {login},
        "password": {password},
        "api_key": api_key,
    }

    try:
        # Выполняем POST-запрос
        response = requests.post(url, data=params)

        # Проверяем статус код
        if response.status_code != 200:
            return {"error": f"HTTP error: {response.status_code}"}

        # Преобразуем ответ в JSON
        json_data = response.json()

        # Проверяем на наличие ошибок в ответе
        if json_data.get("error"):
            return {"error": json_data["error"]}

        return json_data  # Возвращаем данные о регионах

    except Exception as e:
        return {"error": str(e)}


# Функция для отправки одного запроса
def get_price_from_api(login: str, password: str, code: str, brand: str, api_key: str) -> dict | None:
    # URL API
    url = "https://api.zzap.pro/webservice/datasharing.asmx/GetSearchResultLight"

    # Параметры запроса
    params = {
        "login": {login},
        "password": {password},
        "code_region": "",
        "search_text": "",
        "partnumber": {code},
        "class_man": {brand},
        "row_count": "100",
        "type_request": "4",
        "api_key": {api_key}
    }

    # Выполнение GET-запроса
    response = requests.get(url, params=params)

    # Проверка ответа
    if response.status_code == 200:
        try:
            # Парсинг JSON-ответа
            json_data = response.json()
            return json_data

        except ValueError as e:
            print("Ошибка парсинга JSON:", e)
            print(response.text)  # Для диагностики можно вывести сырой текст ответа
            return None
    else:
        print(f"Error: {response.status_code}")
        print(response.text)
        return None


# Функция для извлечения минимальной цены из XML ответа
def extract_min_price_from_response(json_data):
    # Парсим json-ответ
    try:
        min_price = json_data.get('price_min_instock')
        return min_price
    except Exception as e:
        print(f"Ошибка при парсинге JSON: {e}")
        return None


# Основная функция
def process_excel(input_file, interval=5):
    # Загружаем рабочую книгу и активный лист
    try:
        workbook = load_workbook(input_file)
        work_sheet = workbook.active  # Доступ к активному листу
    except Exception as e:
        print(f"Ошибка чтения файла: {e}")
        return

    # Индексы столбцов для "Номенклатура.Производитель" и "Артикул" по названиям
    try:
        brand_column = 1
        code_column = 4
        old_price_column = 10
        price_column = 11
    except ValueError as e:
        print(f"Ошибка: не найдены необходимые столбцы. {e}")
        return

    # Обрабатываем строки (начиная с 11-й строки, т.к. 10-я - это заголовки)
    for row_idx in range(12, work_sheet.max_row + 1):
        brand = work_sheet.cell(row=row_idx, column=brand_column).value
        code = work_sheet.cell(row=row_idx, column=code_column).value
        old_price = work_sheet.cell(row=row_idx, column=old_price_column).value

        if brand is None or code is None:
            continue

        # Выполняем запрос к API
        print(f"Отправляем запрос для артикула: {code}, бренд: {brand}")
        json_data = get_price_from_api(login='pheonix1', password='pPHOENIX11', code=code, brand=brand)

        if json_data:
            # Извлекаем минимальную цену из ответа
            min_price = extract_min_price_from_response(json_data)

            # Если минимальная цена найдена, записываем её в Excel
            if min_price is not None:
                min_price = min_price * 0.8
                work_sheet.cell(row=row_idx, column=price_column).value = min_price
            else:
                work_sheet.cell(row=row_idx,
                                column=price_column).value = old_price  # Если цена не найдена записываем старую цену
        else:
            work_sheet.cell(row=row_idx,
                            column=price_column).value = old_price  # Если данных нет, записываем старую цену

        # Интервал между запросами
        time.sleep(interval)

    # Сохраняем изменения в тот же файл
    try:
        workbook.save(input_file)  # Сохраняем в исходный файл
        print(f"Результаты сохранены в {input_file}")
    except Exception as e:
        print(f"Ошибка записи файла: {e}")


# Точка входа
if __name__ == "__main__":
    # Укажите путь к файлу с артикулами и брендами
    input_file = "data/data.xlsx"  # Входной файл
    interval = 3  # Интервал между запросами в секундах

    process_excel(input_file, interval)
