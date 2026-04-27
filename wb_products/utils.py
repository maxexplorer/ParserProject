import os
from openpyxl import Workbook, load_workbook


def merge_excel_files(input_dir='results', output_file='result_data.xlsx'):

    print(f'Старт: {input_dir}')

    files = [f for f in os.listdir(input_dir) if f.endswith('.xlsx')]

    all_columns = set()

    # ----------------------
    # 1. собираем колонки
    # ----------------------
    for file in files:
        path = os.path.join(input_dir, file)

        try:
            wb = load_workbook(path, read_only=True, data_only=True)

            if 'Data' not in wb.sheetnames:
                continue

            ws = wb['Data']

            headers = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)

            if not headers:
                continue

            headers = [str(h).strip() if h else None for h in headers]

            all_columns.update(
                [h for h in headers if h and not str(h).startswith('Unnamed')]
            )

        except Exception as e:
            print(f'Ошибка колонок {file}: {e}')

    base_columns = ['Название', 'Бренд', 'Цена', 'Размер']
    other_columns = sorted(col for col in all_columns if col not in base_columns)
    final_columns = base_columns + other_columns

    print(f'Всего колонок: {len(final_columns)}')

    # ----------------------
    # 2. output
    # ----------------------
    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet()
    ws_out.append(final_columns)

    col_index = {col: i for i, col in enumerate(final_columns)}

    total_rows = 0

    # ----------------------
    # 3. данные
    # ----------------------
    for file in files:
        path = os.path.join(input_dir, file)

        try:
            print(f'Читаю: {file}')

            wb = load_workbook(path, read_only=True, data_only=True)

            if 'Data' not in wb.sheetnames:
                continue

            ws = wb['Data']

            headers = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)

            if not headers:
                continue

            headers = [str(h).strip() if h else None for h in headers]
            header_map = {h: i for i, h in enumerate(headers) if h}

            rows = ws.iter_rows(min_row=3, values_only=True)

            for row in rows:
                out = [None] * len(final_columns)

                for h, i in header_map.items():
                    if h in col_index and i < len(row):
                        out[col_index[h]] = row[i]

                ws_out.append(out)
                total_rows += 1   # 🔥 ВАЖНО

        except Exception as e:
            print(f'Ошибка {file}: {e}')

    wb_out.save(output_file)

    print(f'Готово: {output_file}')
    print(f'Строк: {total_rows}')


if __name__ == '__main__':
    merge_excel_files()