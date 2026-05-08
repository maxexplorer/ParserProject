# update_quantity.py

import os
import glob
import time

import requests
import openpyxl

from configs.config import API_URLS_OZON, OZON_HEADERS, API_URLS_WB, WB_PRICES_AND_DISCOUNTS_HEADERS, WB_MARKETPLACE_HEADERS
from utils import chunk_list


def get_fbs_quantity_ozon(offer_ids: list[str]) -> dict:
    """
    Получает остатки FBS по списку offer_id.
    Возвращает словарь:
    {
        offer_id: fbs_present,
        ...
    }
    """

    result = {}

    cursor = ''
    limit = 100

    while True:
        payload = {
            'cursor': cursor,
            'filter': {
                'visibility': 'ALL',
                'offer_id': offer_ids
            },
            'limit': limit
        }

        try:
            response = requests.post(
                API_URLS_OZON['product_info_stocks'],
                headers=OZON_HEADERS,
                json=payload,
                timeout=15
            )
            response.raise_for_status()
        except Exception as e:
            print(f'❌ Ошибка API: {e}')
            break

        data = response.json()

        items = data.get('items', [])

        for item in items:
            offer = item.get('offer_id')
            if not offer:
                continue

            stocks = item.get('stocks', [])

            for stock in stocks:
                s_type = stock.get('type', '').lower()
                if s_type == 'fbs':
                    result[offer] = stock.get('present', 0)

        cursor = data.get('cursor')
        if not cursor:
            break

    return result


def get_wb_fbs_warehouses() -> list[dict]:
    """
    Получает список складов продавца WB
    и фильтрует только активные FBS склады
    """

    try:
        response = requests.get(
            API_URLS_WB['warehouses'],
            headers=WB_MARKETPLACE_HEADERS,
            timeout=15
        )
        response.raise_for_status()
    except Exception as ex:
        print(f'❌ Ошибка получения списка складов WB: {ex}')
        return []

    warehouses = response.json()

    warehouse_ids = []

    # Фильтруем только FBS склады
    for warehouse in warehouses:
        if warehouse.get('deliveryType') == 1:
            warehouse_id = warehouse.get('id')
            warehouse_ids.append(warehouse_id)

    if not warehouse_ids:
        print('❌ FBS склады не найдены')
        return []

    return warehouse_ids


def get_size_ids() -> dict[str, list[int]]:
    """
    Возвращает связь:
    {
        vendorCode: [sizeID, sizeID, ...]
    }
    """

    vendor_code_to_size_ids: dict[str, list[int]] = {}

    limit = 1000
    offset = 0

    while True:
        params = {
            'limit': limit,
            'offset': offset
        }

        try:
            time.sleep(1)
            response = requests.get(
                API_URLS_WB['list_goods_filter'],
                headers=WB_PRICES_AND_DISCOUNTS_HEADERS,
                params=params,
                timeout=15
            )
            response.raise_for_status()
        except Exception as ex:
            print(f'❌ Ошибка получения товаров WB: {ex}')
            break

        data = response.json()
        goods = data.get('data', {}).get('listGoods', [])
        if not goods:
            break

        for item in goods:
            vendor_code = item.get('vendorCode')
            if not vendor_code:
                continue

            for size in item.get('sizes', []):
                size_id = size.get('sizeID')
                if not size_id:
                    continue

                vendor_code_to_size_ids.setdefault(vendor_code, []).append(int(size_id))

        if len(goods) < limit:
            break

        offset += limit

    return vendor_code_to_size_ids


def get_fbs_quantity_wb(offer_ids: list[str]) -> dict[str, int]:
    """
    Получает остатки FBS WB по артикулам (vendorCode).

    Возвращает:
    {
        vendorCode: total_amount
    }
    """
    result: dict[str, int] = {vendor: 0 for vendor in offer_ids}

    warehouses_ids = get_wb_fbs_warehouses()
    vendor_to_sizes = get_size_ids()  # vendorCode -> [sizeID]

    for warehouse_id in warehouses_ids:
        print(f'📦 WB FBS склад: {warehouse_id}')

        # Формируем payload для всех артикулов
        all_size_ids = []
        vendor_map = {}  # sizeID -> vendorCode
        for vendor_code in offer_ids:
            sizes = vendor_to_sizes.get(vendor_code, [])
            for size_id in sizes:
                all_size_ids.append(size_id)
                vendor_map[size_id] = vendor_code

        # Разбиваем на чанки по 1000
        for chunk in chunk_list(all_size_ids, 1000):
            payload = {'chrtIds': chunk}

            try:
                response = requests.post(
                    API_URLS_WB['stocks'].format(warehouseId=warehouse_id),
                    headers=WB_MARKETPLACE_HEADERS,
                    json=payload,
                    timeout=15
                )
                response.raise_for_status()
            except Exception as ex:
                print(f'❌ WB stocks error (warehouse={warehouse_id}): {ex}')
                continue

            data = response.json()
            stocks = data.get('stocks', [])

            for stock in stocks:
                size_id = stock.get('chrtId')
                vendor_code = vendor_map.get(size_id)
                if vendor_code:
                    result[vendor_code] += int(stock.get('amount', 0))

    return result


def write_fbs_quantity_to_excel(fbs_data: dict, marketplace='ОЗОН') -> None:
    """
    Записывает остатки FBS в 16-й столбец Excel.
    Старт записи — с 4-й строки.
    """

    folder = 'data'
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    if not excel_files:
        print('❗ В папке data/ не найдено .xlsx файлов.')
        return

    wb = openpyxl.load_workbook(excel_files[0])
    if marketplace not in wb.sheetnames:
        print(f'❗ Лист {marketplace} не найден.')
        return

    ws = wb[marketplace]

    for row in ws.iter_rows(min_row=4):
        offer_id = str(row[1].value).strip()
        if not offer_id:
            continue

        fbs_value = fbs_data.get(offer_id)
        if fbs_value is not None:
            row[15].value = fbs_value  # 16-й столбец (индекс 15)

    wb.save(os.path.join(folder, 'data.xlsx'))
    print('✅ Остатки FBS успешно записаны в Excel (16-й столбец)')


# Для ручного теста
if __name__ == '__main__':
    pass
