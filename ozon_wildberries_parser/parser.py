import os
import re
import time
import json
from random import randint
from urllib.parse import urlparse, urlunparse
from datetime import datetime

import requests
from requests import Session

from new_undetected_chromedriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup

import openpyxl

start_time = datetime.now()

workbook = openpyxl.load_workbook('data/data.xlsm')


def init_undetected_chromedriver():
    """
    Инициализирует undetected_chromedriver, возвращает драйвер
    """
    driver = Chrome()
    driver.maximize_window()
    driver.implicitly_wait(15)
    return driver


def get_cleaned_url(product_url: str) -> str:
    """
    Очищает ссылку Ozon, возвращает product_id
    """
    parsed_url = urlparse(product_url)
    cleaned_url = urlunparse((parsed_url.scheme, parsed_url.netloc, parsed_url.path, '', '', ''))
    product_id = cleaned_url.rstrip('/').split('-')[-1]
    return product_id


def get_products_ids_ozon(driver: Chrome, pages: int, text: str) -> list:
    """
    Парсит Ozon, возвращает список product_ids по запросу text
    """
    products_ids_list = []
    products_ids_set = set()

    try:
        driver.get(url='https://www.ozon.ru/')
        time.sleep(3)

        try:
            input_text = driver.find_element(By.CSS_SELECTOR, 'input[placeholder="Искать на Ozon"]')
            input_text.clear()
            input_text.send_keys(text)
            input_text.send_keys(Keys.ENTER)
            time.sleep(3)
        except Exception as ex:
            print(f'input_text: {ex}')

        for i in range(pages):
            driver.execute_script('window.scrollBy(0, window.innerHeight);')
            time.sleep(randint(2, 3))

            try:
                data_items = driver.find_elements(By.CSS_SELECTOR, 'a[data-prerender="true"].tile-clickable-element')
            except Exception as ex:
                print(f'data_items: {ex}')
                data_items = []

            for item in data_items:
                try:
                    product_url = f'https://www.ozon.ru{item.get_attribute("href")}'
                    product_id = get_cleaned_url(product_url)
                except Exception:
                    continue

                if product_id in products_ids_set:
                    continue

                products_ids_list.append(product_id)
                products_ids_set.add(product_id)

        print(f'Получено: {len(products_ids_set)} ids')
        return products_ids_list

    except Exception as ex:
        print(f'get_products_ids_ozon: {ex}')
        return products_ids_list


def get_products_ids_wb(headers: dict, pages: int, text: str) -> list:
    """
    Парсит Wildberries, возвращает список product_ids по запросу text
    """
    products_ids_list = []

    with Session() as session:
        for page in range(1, pages + 1):
            params = {
                'ab_testing': 'false',
                'appType': '1',
                'curr': 'rub',
                'dest': '-5856841',
                'hide_dtype': '10',
                'lang': 'ru',
                'page': page,
                'query': text,
                'resultset': 'catalog',
                'sort': 'popular',
                'spp': '30',
                'suppressSpellcheck': 'false',
            }

            try:
                time.sleep(1)
                response = session.get(
                    'https://search.wb.ru/exactmatch/ru/common/v9/search',
                    params=params,
                    headers=headers,
                    timeout=60
                )
                if response.status_code != 200:
                    print(f'get_products_ids_wb: status_code {response.status_code}')
            except Exception as ex:
                print(f'get_products_ids_wb: {ex}')
                continue

            try:
                data = response.json()
                for item in data['data']['products']:
                    product_id = item.get('id')
                    products_ids_list.append(str(product_id))
            except Exception as ex:
                print(f'products_ids_wb parse: {ex}')

    print(f'Получено: {len(products_ids_list)} ids')
    return products_ids_list


def ozon_parser(driver: Chrome, workbook: openpyxl.Workbook, pages: int = 3):
    """
    Основной парсер Ozon: записывает позиции, цены, количество и склад
    """
    ws = workbook['ОЗОН']
    processed_texts = {}

    try:
        for row in ws.iter_rows(min_row=4):
            text = row[1].value
            if not text:
                continue

            if text in processed_texts:
                product_ids = processed_texts[text]
            else:
                product_ids = get_products_ids_ozon(driver, pages, text)
                processed_texts[text] = product_ids

            for cell in row:
                if isinstance(cell.value, str) and 'https://' in cell.value:
                    product_url = cell.value
                    try:
                        driver.get(product_url)
                        time.sleep(randint(3, 5))
                        soup = BeautifulSoup(driver.page_source, 'lxml')
                    except Exception:
                        continue

                    try:
                        if soup.find('h2', string=re.compile('Этот товар закончился')):
                            row[cell.column - 4].value = None
                            row[cell.column - 3].value = 0
                            row[cell.column - 2].value = None
                            continue
                    except Exception:
                        pass

                    try:
                        if soup.find('h2', string=re.compile('Такой страницы не существует')):
                            row[cell.column - 4].value = None
                            row[cell.column - 3].value = 'Такой страницы не существует'
                            row[cell.column - 2].value = None
                            continue
                    except Exception:
                        pass

                    try:
                        product_id = get_cleaned_url(product_url)
                        product_position = product_ids.index(product_id) + 1
                    except Exception:
                        product_position = '-'
                    row[cell.column].value = product_position

                    try:
                        price_text = soup.find('span', string=re.compile('c Ozon Картой')).find_parent().text
                        price = ''.join(filter(str.isdigit, price_text))
                    except Exception:
                        price = '-'
                    row[cell.column - 4].value = price

                    try:
                        stor_item = soup.find('span', string=re.compile('Со склада')).text.strip()
                        storage = 'FBO' if 'Со склада Ozon' in stor_item else 'FBS'
                    except Exception:
                        storage = None
                    row[cell.column - 2].value = storage

                    try:
                        add_btn = driver.find_element(By.XPATH, '//div[contains(text(), "Добавить в корзину")]')
                        add_btn.find_element(By.XPATH, '../..').click()
                        time.sleep(randint(3, 5))
                    except Exception:
                        continue

                    try:
                        in_basket_btn = driver.find_element(By.XPATH, '//span[contains(text(), "В корзине")]')
                        in_basket_btn.find_element(By.XPATH, '../../..').click()
                        time.sleep(randint(3, 5))
                    except Exception:
                        continue

                    try:
                        soup = BeautifulSoup(driver.page_source, 'lxml')
                        divs = soup.find_all('div', attrs={'data-state': True})
                        quantity = None
                        for div in divs:
                            try:
                                data_dict = json.loads(div['data-state'])
                                quantity = data_dict['items'][0]['quantity']['maxQuantity']
                                if quantity:
                                    break
                            except:
                                continue
                    except Exception:
                        quantity = None
                    row[cell.column - 3].value = quantity

                    try:
                        share_btn = driver.find_element(By.XPATH, '//div[contains(text(), "Поделиться")]')
                        share_btn.find_element(By.XPATH, '../..').find_element(By.XPATH, 'following-sibling::button').click()
                        WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div[2]/div/div/section/div[3]/button'))
                        ).click()
                        time.sleep(randint(3, 5))
                    except Exception:
                        continue

    except Exception as ex:
        print(ex)
    finally:
        if not os.path.exists('results'):
            os.makedirs('results')
        workbook.save('results/result_data.xlsx')


def wildberries_parser(workbook: openpyxl.Workbook, pages: int = 3):
    """
    Основной парсер Wildberries: записывает позиции, цены, количество и склад
    """
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
    }
    ws = workbook['ВБ']
    processed_texts = {}

    for row in ws.iter_rows(min_row=4):
        text = row[1].value
        if not text:
            continue

        if text in processed_texts:
            product_ids = processed_texts[text]
        else:
            product_ids = get_products_ids_wb(headers, pages, text)
            processed_texts[text] = product_ids

        for cell in row:
            if isinstance(cell.value, str) and 'https://' in cell.value:
                product_url = cell.value
                product_id = product_url.split('catalog')[-1].split('detail')[0].strip('/')

                params = {
                    'appType': '1',
                    'curr': 'rub',
                    'dest': '-1257786',
                    'spp': '30',
                    'nm': {product_id},
                }

                try:
                    time.sleep(randint(1, 3))
                    response = requests.get('https://card.wb.ru/cards/v1/detail', params=params, headers=headers)
                    if response.status_code != 200:
                        continue
                    data = response.json()
                except Exception:
                    continue

                try:
                    quantity = data['data']['products'][0]['sizes'][0]['stocks'][0]['qty']
                    row[cell.column - 3].value = quantity
                except Exception:
                    row[cell.column - 4].value = ''
                    row[cell.column - 3].value = 0
                    row[cell.column - 2].value = ''
                    continue

                try:
                    product_position = product_ids.index(product_id) + 1
                except Exception:
                    product_position = '-'
                row[cell.column].value = product_position

                try:
                    storage_id = data['data']['products'][0]['sizes'][0]['stocks'][0]['dtype']
                    storage = 'FBO' if storage_id == 4 else 'FBS'
                except Exception:
                    storage = None
                row[cell.column - 2].value = storage

                try:
                    price = data['data']['products'][0]['salePriceU'] // 100
                except Exception:
                    price = '-'
                row[cell.column - 4].value = price

    if not os.path.exists('results'):
        os.makedirs('results')
    workbook.save('results/result_data.xlsx')
