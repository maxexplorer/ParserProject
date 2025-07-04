import os
import re
import time
from datetime import datetime
from random import randint
from urllib.parse import urlparse, urlunparse
import json

import requests
from requests import Session

from new_undetected_chromedriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup

import openpyxl

start_time = datetime.now()

# Открываем файл Excel
workbook = openpyxl.load_workbook("data/data.xlsm")


# Функция инициализации объекта chromedriver
def init_undetected_chromedriver():
    driver = Chrome()
    driver.maximize_window()
    driver.implicitly_wait(15)

    return driver


def get_cleaned_url(product_url: str) -> str:
    # Парсим ссылку
    parsed_url = urlparse(product_url)

    # Собираем только основные части: схема, домен и путь
    cleaned_url = urlunparse((parsed_url.scheme, parsed_url.netloc, parsed_url.path, '', '', ''))

    product_id = cleaned_url.rstrip('/').split('-')[-1]

    return product_id


# Функция получения ссылок товаров
def get_products_ids_ozon(driver: Chrome, pages: int, text: str) -> list[str]:
    products_ids_list = list()
    products_ids_set = set()

    try:
        driver.get(url="https://www.ozon.ru/")
        time.sleep(3)

        try:
            input_text = driver.find_element(By.CSS_SELECTOR, 'input[placeholder="Искать на Ozon"]')
            input_text.clear()
            input_text.send_keys(text)
            input_text.send_keys(Keys.ENTER)
            time.sleep(3)
        except Exception as ex:
            print(f'input_text: {ex}')
        #
        # try:
        #     search_product = driver.find_element(By.CSS_SELECTOR, 'button[aria-label="Поиск"]')
        #     search_product.click()
        #     time.sleep(3)
        # except Exception as ex:
        #     print(f'search_product: {ex}')

        for i in range(pages):
            # driver.execute_script("window.scrollTo(0, 4000);")
            driver.execute_script("window.scrollBy(0, window.innerHeight);")

            time.sleep(randint(2, 3))

            try:
                data_items = driver.find_elements(By.CSS_SELECTOR, 'a[data-prerender="true"].tile-clickable-element')

            except Exception as ex:
                print(f'data_items: - {ex}')
                data_items = []

            for item in data_items:
                try:
                    product_url = f"https://www.ozon.ru{item.get_attribute('href')}"

                    product_id = get_cleaned_url(product_url=product_url)

                except Exception:
                    continue

                if product_id in products_ids_set:
                    continue

                products_ids_list.append(product_id)
                products_ids_set.add(product_id)

        print(f'Получено: {len(products_ids_set)} ids')

        # if not os.path.exists('data'):
        #     os.makedirs('data')
        #
        # with open(f'data/products_ids_list_ozon.txt', 'a', encoding='utf-8') as file:
        #     print(*products_ids_list, file=file, sep='\n')

        return products_ids_list

    except Exception as ex:
        print(f'get_products_urls: {ex}')


# Функция получения ссылок товаров
def get_products_ids_wb(headers: dict, pages: int, text: str) -> list[str]:
    products_ids_list = list()

    with Session() as session:
        for page in range(1, pages + 1):
            params = {
                'ab_testing': 'false',
                'appType': '1',
                'curr': 'rub',
                'dest': '-5856841',
                'hide_dtype': '10',
                'lang': 'ru',
                'page': page,
                'query': text,
                'resultset': 'catalog',
                'sort': 'popular',
                'spp': '30',
                'suppressSpellcheck': 'false',
            }

            try:
                time.sleep(1)
                response = session.get(
                    f'https://search.wb.ru/exactmatch/ru/common/v9/search',
                    params=params,
                    headers=headers,
                    timeout=60
                )

                if response.status_code != 200:
                    print(f'get_products_ids_wb: status_code {response.status_code}')

            except Exception as ex:
                print(f'get_products_ids_wb: {ex}')
                continue

            data = response.json()

            try:
                for item in data['data']['products']:
                    product_id = item.get('id')
                    products_ids_list.append(str(product_id))
            except Exception as ex:
                print(f'products_ids: {ex}')

    print(f'Получено: {len(products_ids_list)} ids')

    # if not os.path.exists('data'):
    #     os.makedirs('data')
    #
    # with open(f'data/products_ids_list_wb.txt', 'a', encoding='utf-8') as file:
    #     print(*products_ids_list, file=file, sep='\n')

    return products_ids_list


def ozon_parser(driver: Chrome, workbook: openpyxl.Workbook, pages: int = 3):
    # Выбираем активный лист (или любой другой лист)
    ws = workbook['ОЗОН']

    # Словарь для хранения уже обработанных текстов и их product_ids
    processed_texts = {}

    try:
        for row in ws.iter_rows(min_row=4):
            text = row[1].value

            if not text:
                continue

            # Если text уже обработан, берем product_ids из словаря, иначе вызываем функцию
            if text in processed_texts:
                product_ids = processed_texts[text]
            else:
                product_ids = get_products_ids_ozon(driver=driver, pages=pages, text=text)
                processed_texts[text] = product_ids

            for cell in row:
                # Проверяем, что ячейка содержит строку
                if isinstance(cell.value, str) and 'https://' in cell.value:
                    try:
                        product_url = cell.value

                        driver.get(url=product_url)

                        time.sleep(randint(3, 5))

                        soup = BeautifulSoup(driver.page_source, 'lxml')
                    except Exception as ex:
                        continue

                    try:
                        out_of_stock = soup.find('h2', string=re.compile('Этот товар закончился'))
                        if out_of_stock:
                            # print(f'{url}: Этот товар закончился')
                            row[cell.column - 4].value = None
                            row[cell.column - 3].value = 0
                            row[cell.column - 2].value = None
                            continue
                    except Exception:
                        pass

                    try:
                        no_such_page = soup.find('h2', string=re.compile('Такой страницы не существует'))
                        if no_such_page:
                            row[cell.column - 4].value = None
                            row[cell.column - 3].value = 'Такой страницы не существует'
                            row[cell.column - 2].value = None
                            print(f'{product_url}: Такой страницы не существует')
                            continue
                    except Exception:
                        pass

                    try:
                        product_id = get_cleaned_url(product_url=product_url)
                    except Exception as ex:
                        product_id = None

                    try:
                        product_position = product_ids.index(product_id) + 1
                    except Exception as ex:
                        # print(f'product_position: {product_url} - {ex}')
                        product_position = '-'
                    row[cell.column].value = product_position

                    try:
                        price = ''.join(filter(lambda x: x.isdigit(), soup.find('span', string=re.compile(
                            'c Ozon Картой')).find_parent().text))
                    except Exception as ex:
                        # print(f'price - {ex}')
                        price = '-'
                    row[cell.column - 4].value = price

                    try:
                        stor_item = soup.find('span', string=re.compile('Со склада')).text.strip()
                        storage = 'FBO' if 'Со склада Ozon' in stor_item else 'FBS'

                    except Exception as ex:
                        # print(f'storage - {ex}')
                        storage = None
                    row[cell.column - 2].value = storage

                    try:
                        add_in_basket = driver.find_element(By.XPATH, '//div[contains(text(), "Добавить в корзину")]')
                        parent_element = add_in_basket.find_element(By.XPATH, "../..")
                        parent_element.click()
                        time.sleep(randint(3, 5))
                    except Exception as ex:
                        # print(f'add_in_basket: {ex}')
                        continue

                    try:
                        in_basket = driver.find_element(By.XPATH, '//span[contains(text(), "В корзине")]')
                        parent_element = in_basket.find_element(By.XPATH, "../../..")
                        parent_element.click()
                        time.sleep(randint(3, 5))

                    except Exception as ex:
                        # print(f'in_basket: {ex}')
                        continue

                    try:
                        soup = BeautifulSoup(driver.page_source, 'lxml')

                        divs = soup.find_all('div', attrs={'data-state': True})

                        quantity = None

                        for div in divs:
                            data_state = div['data-state']
                            data_dict = json.loads(data_state)
                            try:
                                quantity = data_dict['items'][0]['quantity']['maxQuantity']
                                if quantity:
                                    break
                            except (KeyError, IndexError, TypeError):
                                continue
                    except Exception as ex:
                        # print(f'quantity - {ex}')
                        quantity = None
                    row[cell.column - 3].value = quantity

                    try:
                        button_share = driver.find_element(By.XPATH, '//div[contains(text(), "Поделиться")]')
                        parent_element = button_share.find_element(By.XPATH, "../..")
                        button_del1 = parent_element.find_element(By.XPATH, 'following-sibling::button')
                        button_del1.click()
                    except Exception as ex:
                        # print(f'button_del1: {ex}')
                        continue

                    try:
                        button_del2 = WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable(
                                (By.XPATH, '/html/body/div[4]/div/div[2]/div/div/section/div[3]/button'))
                        )
                        button_del2.click()

                        time.sleep(randint(3, 5))
                    except Exception as ex:
                        # print(f'button_del2: {ex}')
                        continue

                    print(f'{product_url}: position - {product_position}, price - {price}, quantity - {quantity}, '
                          f'storage - {storage}')

    except Exception as ex:
        print(ex)
    finally:
        if not os.path.exists('results'):
            os.makedirs('results')

        workbook.save('results/result_data.xlsx')


def wildberries_parser(workbook: openpyxl.Workbook, pages: int = 3):
    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'max-age=0',
        # 'cookie': '_wbauid=1430232901729079545; wbx-validation-key=042261a8-d7e3-4266-8343-31fb35d5a295',
        'priority': 'u=0, i',
        'referer': 'https://www.wildberries.ru/',
        'sec-ch-ua': '"Not A(Brand";v="8", "Chromium";v="132", "Google Chrome";v="132"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
    }

    # Выбираем активный лист (или любой другой лист)
    ws = workbook['ВБ']

    # Словарь для хранения уже обработанных текстов и их product_ids
    processed_texts = {}

    for row in ws.iter_rows(min_row=4):
        text = row[1].value

        if not text:
            continue

        # Если text уже обработан, берем product_ids из словаря, иначе вызываем функцию
        if text in processed_texts:
            product_ids = processed_texts[text]
        else:
            product_ids = get_products_ids_wb(headers=headers, pages=pages, text=text)
            processed_texts[text] = product_ids

        for cell in row:
            # Проверяем, что ячейка содержит строку
            if isinstance(cell.value, str) and 'https://' in cell.value:
                product_url = cell.value
                product_id = product_url.split('catalog')[-1].split('detail')[0].strip('/')

                params = {
                    'appType': '1',
                    'curr': 'rub',
                    'dest': '-1257786',
                    'spp': '30',
                    'nm': {product_id},
                }

                try:
                    time.sleep(randint(1, 3))
                    response = requests.get('https://card.wb.ru/cards/v1/detail', params=params, headers=headers)
                    if response.status_code != 200:
                        print(f'{product_url}: {response.status_code}')
                except Exception as ex:
                    print(f'{product_url}: ex')
                    continue

                data = response.json()

                try:
                    quantity = data['data']['products'][0]['sizes'][0]['stocks'][0]['qty']
                    row[cell.column - 3].value = quantity
                except Exception:
                    print(f'{product_url}: Этот товар закончился')
                    row[cell.column - 4].value = ''
                    row[cell.column - 3].value = 0
                    row[cell.column - 2].value = ''
                    continue

                try:
                    product_position = product_ids.index(product_id) + 1
                except Exception as ex:
                    # print(f'product_position: {product_url} - {ex}')
                    product_position = '-'

                row[cell.column].value = product_position

                try:
                    storage_id = data['data']['products'][0]['sizes'][0]['stocks'][0]['dtype']
                    storage = 'FBO' if storage_id == 4 else 'FBS'
                except Exception:
                    storage = None
                row[cell.column - 2].value = storage

                try:
                    price = data['data']['products'][0]['salePriceU'] // 100
                except Exception:
                    price = '-'
                row[cell.column - 4].value = price

                print(
                    f'{product_url}: position - {product_position}, price - {price}, quantity - {quantity}, storage - {storage}')

    if not os.path.exists('results'):
        os.makedirs('results')

    workbook.save('results/result_data.xlsx')


def main():
    try:
        value = input('Введите значение:\n1 - Ozon\n2 - Wildberries\n3 - Оба сайта\n')
    except Exception:
        raise 'Введено неправильное значение:\n'

    if value == '1':
        pages = int(input('Введите количество страниц Ozon: \n'))
        driver = init_undetected_chromedriver()
        try:
            print('Сбор данных Ozon')
            ozon_parser(driver=driver, workbook=workbook, pages=pages)
            print('Сбор данных Ozon завершен')
        except Exception as ex:
            print(f'main: {ex}')
            input("Нажмите Enter, чтобы закрыть программу...")
        finally:
            driver.close()
            driver.quit()
    elif value == '2':
        try:
            pages = int(input('Введите количество страниц Wildberries: \n'))
            print('Сбор данных Wildberries')
            wildberries_parser(workbook=workbook, pages=pages)
            print('Сбор данных Wildberries завершен')
        except Exception as ex:
            print(f'main: {ex}')
            input("Нажмите Enter, чтобы закрыть программу...")
    elif value == '3':
        pages_ozon = int(input('Введите количество страниц Ozon: \n'))
        pages_wb = int(input('Введите количество страниц Wildberries: \n'))
        driver = init_undetected_chromedriver()
        try:
            print('Сбор данных Ozon')
            ozon_parser(driver=driver, workbook=workbook, pages=pages_ozon)
            print('Сбор данных Ozon завершен')
        except Exception as ex:
            print(f'main: {ex}')
            input("Нажмите Enter, чтобы закрыть программу...")
        finally:
            driver.close()
            driver.quit()
        try:
            print('Сбор данных Wildberries')
            wildberries_parser(workbook=workbook, pages=pages_wb)
            print('Сбор данных Wildberries завершен')
        except Exception as ex:
            print(f'main: {ex}')
            input("Нажмите Enter, чтобы закрыть программу...")
    else:
        print('Введено неправильное значение')

    execution_time = datetime.now() - start_time
    print(f'Время работы программы: {execution_time}')


if __name__ == '__main__':
    main()
