# update_volume.py

import os
import glob

import pandas as pd
import requests
import openpyxl

from configs.config import API_URLS_OZON, OZON_HEADERS, WB_HEADERS


def load_offer_id_from_excel(sheet_name: str) -> list[str]:
    """
    Загружает offer_id из Excel (2-й столбец), начиная с 3 строки (skiprows=2).
    Возвращает список offer_id.
    """
    folder = 'data'
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    if not excel_files:
        print('❗ В папке data/ не найдено .xlsx файлов.')
        return []

    # читаем Excel
    df = pd.read_excel(excel_files[0], sheet_name=sheet_name, skiprows=2)
    df.columns = df.columns.str.strip()

    offer_ids = []

    for _, row in df.iterrows():
        offer_id = str(row.iloc[1]).strip()  # 2-й столбец
        if offer_id and offer_id.lower() != 'nan':
            offer_ids.append(offer_id)

    return offer_ids


def get_volume_ozon(offer_ids: list[str]) -> dict:
    """
    Запрашивает габариты товаров по API v4/product/info/attributes
    и вычисляет объем (м³) по формуле height * depth * width / 1_000_000.
    Возвращает словарь {offer_id: volume_m3}
    """
    result = {}

    data = {
        'filter': {
            'offer_id': offer_ids,
            'visibility': 'ALL'
        },
        'limit': len(offer_ids),
        'sort_dir': 'ASC'
    }

    try:
        response = requests.post(
            API_URLS_OZON['product_info_attributes'],
            headers=OZON_HEADERS,
            json=data,
            timeout=20
        )
        response.raise_for_status()
    except Exception as ex:
        print(f"❌ Ошибка API Ozon при получении габаритов: {ex}")
        return result

    resp = response.json()
    items = resp.get('result', [])

    for item in items:
        offer_id = item.get('offer_id')
        if not offer_id:
            continue

        height = item.get('height', 0)
        depth = item.get('depth', 0)
        width = item.get('width', 0)

        try:
            volume = (height * depth * width) / 1_000_000  # м³
        except:
            volume = 0

        result[offer_id] = round(volume, 5)

    return result


def write_volume_to_excel(volume_data: dict, marketplace='ОЗОН') -> None:
    """
    Записывает объем (м³) в 15-й столбец Excel.
    Старт записи — с 4-й строки (как и цены).
    """
    folder = 'data'
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    if not excel_files:
        print('❗ В папке data/ не найдено .xlsx файлов.')
        return

    wb = openpyxl.load_workbook(excel_files[0])
    if marketplace not in wb.sheetnames:
        print(f'❗ Лист {marketplace} не найден.')
        return

    ws = wb[marketplace]

    for row in ws.iter_rows(min_row=4):
        offer_id = str(row[1].value).strip()
        if not offer_id:
            continue

        volume_value = volume_data.get(offer_id)
        if volume_value is not None:
            row[14].value = round(volume_value, 2)      # 15-й столбец (индекс 14)

    wb.save(os.path.join(folder, 'data.xlsx'))
    print("✅ Объемы успешно записаны в Excel (15-й столбец)")
