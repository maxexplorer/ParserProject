# update_price.py

import os
import glob
import time

import pandas as pd
import requests
import openpyxl

from configs.config import API_URLS_OZON, API_URLS_WB, OZON_HEADERS, WB_HEADERS


def load_article_info_from_excel(sheet_name: str) -> dict:
    """
    Загружает артикулы и new_price из Excel, начиная с 3 строки (skiprows=2).
    Артикул в 1 столбце, new_price в 6 столбце.
    Возвращает словарь {offer_id: delta}
    """
    folder = 'data'
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    if not excel_files:
        print('❗ В папке data/ не найдено .xlsx файлов.')
        return {}

    df = pd.read_excel(excel_files[0], sheet_name=sheet_name, skiprows=2)
    df.columns = df.columns.str.strip()

    article_info = {}
    for _, row in df.iterrows():
        offer_id = str(row.iloc[0]).strip()
        if not offer_id:
            continue

        new_price = row.iloc[5]  # 6 столбец

        if pd.isna(new_price) or new_price == '':
            continue

        try:
            delta = float(new_price)
            article_info[offer_id] = delta
        except Exception as ex:
            print(f'Ошибка преобразования для {offer_id}: {ex}')
            continue

    return article_info


def write_price_to_excel(current_prices: dict, marketplace='ОЗОН') -> None:
    """
    Обновляет цены в 5-м столбце (индекс 4) Excel для указанного листа marketplace,
    сохраняя изменения в data/data.xlsx.
    """
    folder = 'data'
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    if not excel_files:
        print('❗ В папке data/ не найдено .xlsx файлов.')
        return

    wb = openpyxl.load_workbook(excel_files[0])
    if marketplace not in wb.sheetnames:
        print(f'❗ В книге нет листа "{marketplace}"')
        return

    ws = wb[marketplace]

    for row in ws.iter_rows(min_row=4):
        cell_article = row[0].value
        if cell_article:
            offer_id = str(cell_article).strip()
            price_value = current_prices.get(offer_id)
            if price_value is not None:
                row[4].value = price_value  # 5-й столбец

    if not os.path.exists(folder):
        os.makedirs(folder)

    output_path = os.path.join(folder, 'data.xlsx')
    wb.save(output_path)
    print(f'✅ Текущие цены успешно записаны в {output_path}')


def get_current_prices_ozon() -> dict:
    """
    Получает текущие маркетинговые цены товаров с Ozon.
    :return: Словарь {offer_id: current_price}
    """
    result = {}
    cursor = ''
    limit = 100

    while True:
        data = {
            'cursor': cursor,
            'filter': {
                'offer_id': [],
                'product_id': [],
                'visibility': 'IN_SALE'
            },
            'limit': limit
        }

        try:
            time.sleep(1)
            response = requests.post(
                API_URLS_OZON['product_info_prices'],
                headers=OZON_HEADERS,
                json=data,
                timeout=15
            )
            response.raise_for_status()
        except Exception as ex:
            print(f'❌ Ошибка получения цен Ozon: {ex}')
            break

        try:
            data = response.json()
        except ValueError:
            print(f'❌ Ошибка при декодировании JSON.')
            break

        items = data.get('items', [])
        for item in items:
            offer_id = item.get('offer_id')
            price_info = item.get('price', {})
            price = price_info.get('price', 0)
            if offer_id:
                result[offer_id] = float(price)

        cursor = data.get('last_id')
        if not cursor:
            break

    return result


def get_current_prices_wb() -> tuple[dict, dict]:
    vendor_code_to_price = {}
    vendor_code_to_nmID = {}
    limit = 1000
    offset = 0

    while True:
        params = {
            'order': 'nmId',
            'direction': 'asc',
            'limit': limit,
            'offset': offset
        }

        try:
            time.sleep(1)
            response = requests.get(
                API_URLS_WB['list_goods_filter'],
                headers=WB_HEADERS,
                params=params,
                timeout=15
            )
            response.raise_for_status()
        except Exception as ex:
            print(f'❌ Ошибка получения цен WB: {ex}')
            break

        try:
            data = response.json()
        except ValueError:
            print(f'❌ Ошибка при декодировании JSON.')
            break

        goods = data.get('data', {}).get('listGoods', [])
        if not goods:
            break

        for item in goods:
            nm_id = item.get('nmID')
            vendor_code = item.get('vendorCode')
            sizes = item.get('sizes', [])
            if sizes and vendor_code:
                price = sizes[0].get('price', 0)
                vendor_code_to_price[vendor_code] = float(price)
                vendor_code_to_nmID[vendor_code] = nm_id

        offset += limit

    return vendor_code_to_price, vendor_code_to_nmID


def update_prices_ozon(article_info: dict) -> dict:
    """
    Обновляет цены на Ozon по API и возвращает текущие цены для записи в Excel.
    """
    current_prices = get_current_prices_ozon()
    prices = {'prices': []}

    for offer_id, delta in article_info.items():
        current_price = current_prices.get(offer_id)
        if current_price is None:
            print(f'❗ Не найдена текущая цена для Ozon offer_id: {offer_id}')
            continue

        new_price = current_price + delta

        prices['prices'].append({
            'auto_action_enabled': 'UNKNOWN',
            'auto_add_to_ozon_actions_list_enabled': 'UNKNOWN',
            'currency_code': 'RUB',
            'min_price': str(int(new_price)),
            'min_price_for_auto_actions_enabled': True,
            'net_price': '0',
            'offer_id': offer_id,
            'old_price': '0',
            'price': str(int(new_price)),
            'price_strategy_enabled': 'UNKNOWN',
            'product_id': 0,
            'quant_size': 1,
            'vat': '0.1'
        })

    if prices['prices']:
        try:
            response = requests.post(
                API_URLS_OZON['import_price'],
                headers=OZON_HEADERS,
                json=prices,
                timeout=20
            )
        except Exception as ex:
            print(f'❌ Ошибка при обновлении цен Ozon: {ex}')

    return current_prices


def update_prices_wb(article_info: dict) -> dict:
    """
    Обновляет цены на Ozon по API и возвращает текущие цены для записи в Excel.
    """
    vendor_code_to_price, vendor_code_to_nmID = get_current_prices_wb()
    data = {'data': []}

    for vendor_code, delta in article_info.items():
        current_price = vendor_code_to_price.get(vendor_code)
        nm_id = vendor_code_to_nmID.get(vendor_code)

        if current_price is None:
            print(f'❗ Не найдена текущая цена для WB vendorCode: {vendor_code}')
            continue
        if nm_id is None:
            print(f'❗ Не найден nmID для WB vendorCode: {vendor_code}')
            continue

        new_price = current_price + delta

        data['data'].append({
            'nmID': int(nm_id),
            'price': int(new_price),
            'discount': 30
        })

    if data['data']:
        try:
            response = requests.post(
                API_URLS_WB['upload_task'],
                headers=WB_HEADERS,
                json=data,
                timeout=20
            )
            response.raise_for_status()
        except Exception as ex:
            print(f'❌ Ошибка при обновлении цен WB: {ex}')

    return vendor_code_to_price
