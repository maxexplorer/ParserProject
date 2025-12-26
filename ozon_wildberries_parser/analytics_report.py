# analytics_report.py

import os
import glob
from datetime import datetime, timedelta

import requests
import pandas as pd
import openpyxl

from configs.config import API_URLS_OZON, OZON_HEADERS, API_URLS_WB, WB_ANALYTICS_HEADERS


def load_sku_article_from_excel(sheet_name: str) -> dict:
    """
    Загружает артикулы и new_price из Excel, начиная с 3 строки (skiprows=2).
    Артикул в 1 столбце, new_price в 6 столбце.
    Возвращает словарь {offer_id: delta}
    """

    folder = 'data'
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    if not excel_files:
        print('❗ В папке data/ не найдено .xlsx файлов.')
        return {}

    df = pd.read_excel(excel_files[0], sheet_name=sheet_name, skiprows=2)
    df.columns = df.columns.str.strip()

    article_info = {}
    for _, row in df.iterrows():
        offer_id = str(row.iloc[1]).strip()
        if not offer_id:
            continue

        value = row.iloc[0]
        if pd.notna(value):
            if isinstance(value, float) and value.is_integer():
                sku = str(int(value)).strip()
            else:
                sku = str(value).strip()
        else:
            sku = ''

        if pd.isna(offer_id) or offer_id == '' or pd.isna(sku) or sku == '':
            continue

        if sku and offer_id:
            article_info[sku] = offer_id

    return article_info


def get_ozon_orders_report(
        period: str,
        product_to_offer: dict,
        custom_date_from: str = None,
        custom_date_to: str = None
) -> dict:
    """
    Получение количества заказов по offer_id за месяц, неделю или произвольный период.
    product_to_offer — словарь {product_id: offer_id}.
    Для period='custom' необходимо передавать custom_date_from и custom_date_to в формате 'YYYY-MM-DD'.
    """

    if period == 'month':
        date_from = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        date_to = datetime.now().strftime('%Y-%m-%d')
    elif period == 'week':
        date_from = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        date_to = datetime.now().strftime('%Y-%m-%d')
    elif period == 'custom':
        if not custom_date_from or not custom_date_to:
            raise ValueError('Для period="custom" необходимо передать custom_date_from и custom_date_to.')
        date_from = custom_date_from
        date_to = custom_date_to
    else:
        raise ValueError('period должен быть "month", "week" или "custom"')

    data = {
        'date_from': date_from,
        'date_to': date_to,
        'metrics': ['ordered_units'],
        'dimension': ['sku', period],
        'filters': [],
        'limit': 1000,
        'offset': 0
    }

    orders = {}
    while True:
        try:
            response = requests.post(
                API_URLS_OZON['analytics_data'],
                headers=OZON_HEADERS,
                json=data,
                timeout=20
            )
            response.raise_for_status()
            data_dict = response.json()

            for item in data_dict['result']['data']:
                product_id = str(item['dimensions'][0]['id'])
                ordered_units = int(item['metrics'][0])

                offer_id = product_to_offer.get(product_id)
                orders[offer_id] = orders.get(offer_id, 0) + ordered_units

            if len(data_dict['result']['data']) < data['limit']:
                break
            data['offset'] += data['limit']

        except Exception as ex:
            print(f'❌ Ошибка получения данных OZON: {ex}')
            break

    return orders


# analytics_report.py (часть для WB)

import requests
from datetime import datetime, timedelta
from configs.config import API_URLS_WB, WB_ANALYTICS_HEADERS


def get_wb_orders_report(period: str, custom_date_from: str = None, custom_date_to: str = None) -> dict:
    """
    Получение количества заказов по vendorCode за месяц, неделю или произвольный период.
    Использует новый API WB: /analytics/v3/sales-funnel/products
    Возвращает словарь {vendorCode: orderCount}.
    """

    now = datetime.now()

    # Определяем даты
    if period == 'month':
        date_from = (now - timedelta(days=30)).strftime('%Y-%m-%d')
        date_to = now.strftime('%Y-%m-%d')
    elif period == 'week':
        date_from = (now - timedelta(days=7)).strftime('%Y-%m-%d')
        date_to = now.strftime('%Y-%m-%d')
    elif period == 'custom':
        if not custom_date_from or not custom_date_to:
            raise ValueError('Для period="custom" необходимо передать custom_date_from и custom_date_to.')
        date_from = custom_date_from
        date_to = custom_date_to
    else:
        raise ValueError('period должен быть "month", "week" или "custom"')

    page = 1
    limit = 100
    orders: dict[str, int] = {}

    while True:
        payload = {
            "selectedPeriod": {
                "start": date_from,
                "end": date_to
            },
            "nmIds": [],  # можно оставить пустым для всех товаров
            "skipDeletedNm": True,
            "limit": limit,
            "offset": (page - 1) * limit
        }

        try:
            response = requests.post(
                API_URLS_WB['products'],
                headers=WB_ANALYTICS_HEADERS,
                json=payload,
                timeout=20
            )
            response.raise_for_status()
            data = response.json()
        except Exception as ex:
            print(f'❌ Ошибка получения данных WB: {ex}')
            break

        products = data.get('data', {}).get('products', [])
        if not products:
            break

        for product_entry in products:
            product = product_entry.get('product', {})
            statistic = product_entry.get('statistic', {})
            selected = statistic.get('selected', {})
            order_count = selected.get('orderCount', 0)

            vendor_code = product.get('vendorCode')
            if vendor_code:
                orders[vendor_code] = orders.get(vendor_code, 0) + order_count

        # Проверяем, если меньше limit, значит последняя страница
        if len(products) < limit:
            break
        page += 1

    return orders


def write_analytics_to_excel(
        analytics_data: dict,
        marketplace='ОЗОН',
        column_month=12,
        column_week=13,
        column_custom=14,
        period='month'
) -> None:
    """
    Записывает данные аналитики (кол-во заказов) в Excel в указанный столбец по period.

    analytics_data: {offer_id: orders_count}
    period: 'month', 'week' или 'custom' для определения в какую колонку писать
    """
    folder = 'data'
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    if not excel_files:
        print('❗ В папке data/ не найдено .xlsx файлов.')
        return

    wb = openpyxl.load_workbook(excel_files[0])
    if marketplace not in wb.sheetnames:
        print(f'❗ В книге нет листа "{marketplace}"')
        return

    ws = wb[marketplace]

    if period == 'month':
        target_column = column_month
    elif period == 'week':
        target_column = column_week
    elif period == 'custom':
        target_column = column_custom
    else:
        print(f'❌ Неверное значение period: {period}. Должно быть "month", "week" или "custom".')
        return

    for row in ws.iter_rows(min_row=4):
        cell_article = row[1].value
        if not cell_article:
            continue
        offer_id = str(cell_article).strip()
        if offer_id in analytics_data:
            orders_count = analytics_data[offer_id]
            row[target_column - 1].value = orders_count

    if not os.path.exists(folder):
        os.makedirs(folder)

    output_path = os.path.join(folder, 'data.xlsx')
    wb.save(output_path)
    print(f'✅ Данные аналитики успешно записаны в {output_path}')
