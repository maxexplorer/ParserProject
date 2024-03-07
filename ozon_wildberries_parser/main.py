import json
import re
import time
from datetime import datetime
from random import randint

from undetected_chromedriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup

import openpyxl

import requests

start_time = datetime.now()

# Открываем файл Excel
workbook = openpyxl.load_workbook("data/data.xlsx")


def ozone_parser(workbook):
    # Выбираем активный лист (или любой другой лист)
    ws = workbook['ОЗОН']

    driver = Chrome()
    driver.maximize_window()
    driver.implicitly_wait(15)

    try:
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if cell.hyperlink is not None:
                    try:
                        url = cell.hyperlink.target
                        driver.get(url=url)
                        time.sleep(randint(3, 5))

                        soup = BeautifulSoup(driver.page_source, 'lxml')
                    except Exception as ex:
                        continue

                    try:
                        out_of_stock = soup.find('h2', string=re.compile('Этот товар закончился'))
                        if out_of_stock:
                            row[cell.column - 4].value = ''
                            row[cell.column - 3].value = 'Этот товар закончился'
                            row[cell.column - 2].value = ''
                            print(f'{cell.hyperlink.target}: Этот товар закончился')
                            continue
                    except Exception:
                        pass

                    try:
                        no_such_page = soup.find('h2', string=re.compile('Такой страницы не существует'))
                        if no_such_page:
                            row[cell.column - 4].value = ''
                            row[cell.column - 3].value = 'Такой страницы не существует'
                            row[cell.column - 2].value = ''
                            print(f'{cell.hyperlink.target}: Такой страницы не существует')
                            continue
                    except Exception:
                        pass

                    try:
                        price = ''.join(filter(lambda x: x.isdigit(), soup.find('span', string=re.compile(
                            'c Ozon Картой')).find_parent().text))
                    except Exception as ex:
                        # print(f'price - {ex}')
                        price = ''
                    row[cell.column - 4].value = price

                    try:
                        stor_item = soup.find('span', string=re.compile('Со склада')).text.strip()
                        storage = 'FBO' if 'Со склада Ozon' in stor_item else 'FBS'

                    except Exception as ex:
                        # print(f'storage - {ex}')
                        storage = ''
                    row[cell.column - 2].value = storage

                    try:
                        add_in_basket = driver.find_element(By.XPATH, '//div[contains(text(), "Добавить в корзину")]')
                        parent_element = add_in_basket.find_element(By.XPATH, "../..")
                        parent_element.click()
                        time.sleep(randint(3, 5))
                    except Exception as ex:
                        # print(f'add_in_basket: {ex}')
                        continue

                    try:
                        in_basket = driver.find_element(By.XPATH, '//span[contains(text(), "Корзина")]')
                        in_basket.click()
                        time.sleep(randint(3, 5))

                    except Exception as ex:
                        # print(f'in_basket: {ex}')
                        continue

                    try:
                        soup = BeautifulSoup(driver.page_source, 'lxml')
                        items = soup.find_all('input', inputmode='numeric')
                        quantity = items[0].get('max')
                    except Exception as ex:
                        # print(f'quantity - {ex}')
                        quantity = ''
                    row[cell.column - 3].value = quantity

                    try:
                        button_del1 = driver.find_element(By.XPATH,
                                                          '//*[@id="layoutPage"]/div[1]/div/div/div[2]/div[4]/div[1]/div/div/div[1]/div[1]/button')
                        button_del1.click()
                    except Exception as ex:
                        # print(f'button_del1: {ex}')
                        continue

                    try:
                        button_del2 = WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable(
                                (By.XPATH, '/html/body/div[3]/div/div[2]/div/div/section/div[3]/button'))
                        )
                        button_del2.click()

                        time.sleep(randint(3, 5))
                    except Exception as ex:
                        # print(f'button_del2: {ex}')
                        continue

                    print(f'{cell.hyperlink.target}: price - {price}, quantity - {quantity}, storage - {storage}')

    except Exception as ex:
        print(ex)
    finally:
        workbook.save('data/result_data.xlsx')
        driver.close()
        driver.quit()


def wildberries_parser(workbook):
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'Connection': 'keep-alive',
        'Origin': 'https://www.wildberries.ru',
        'Referer': 'https://www.wildberries.ru/catalog/18172488/detail.aspx',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'cross-site',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                      ' (KHTML like Gecko) Chrome/51.0.2704.79 Safari/537.36 Edge/14.14931',
        'sec-ch-ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    # Выбираем активный лист (или любой другой лист)
    ws = workbook['ВБ']

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            if cell.hyperlink is not None:
                url = cell.hyperlink.target
                prod_id = url.split('catalog')[-1].split('detail')[0].strip('/')

                params = {
                    'appType': '1',
                    'curr': 'rub',
                    'dest': '-1257786',
                    'spp': '30',
                    'nm': {prod_id},
                }

                try:
                    time.sleep(randint(1, 3))
                    response = requests.get('https://card.wb.ru/cards/v1/detail', params=params, headers=headers)
                    if response.status_code != 200:
                        print(f'{url}: {response.status_code}')
                except Exception as ex:
                    print(f'{url}: ex')
                    continue

                data = response.json()

                try:
                    quantity = data['data']['products'][0]['sizes'][0]['stocks'][0]['qty']
                    row[cell.column - 3].value = quantity
                except Exception:
                    print(f'{url}: Этот товар закончился')
                    row[cell.column - 4].value = ''
                    row[cell.column - 3].value = ''
                    row[cell.column - 2].value = 'Этот товар закончился'
                    continue

                try:
                    storage_id = data['data']['products'][0]['sizes'][0]['stocks'][0]['dtype']
                    storage = 'FBO' if storage_id == 4 else 'FBS'
                except Exception:
                    storage = None
                row[cell.column - 2].value = storage

                try:
                    price = str(data['data']['products'][0]['salePriceU']).rstrip('00')
                except Exception:
                    price = None
                row[cell.column - 4].value = price

                print(f'{url}: price - {price}, quantity - {quantity}, storage - {storage}')

    workbook.save('data/result_data.xlsx')


def main():
    value = input('Введите значение:\n1 - Ozone\n2 - Wildberries\n3 - Оба сайта\n')

    if value == '1':
        print('Сбор данных Ozone')
        ozone_parser(workbook=workbook)
        print('Сбор данных Ozone завершен')
    elif value == '2':
        print('Сбор данных Wildberries')
        wildberries_parser(workbook=workbook)
        print('Сбор данных Wildberries завершен')
    elif value == '3':
        print('Сбор данных Ozone')
        ozone_parser(workbook=workbook)
        print('Сбор данных Ozone завершен')

        print('Сбор данных Wildberries')
        wildberries_parser(workbook=workbook)
        print('Сбор данных Wildberries завершен')
    else:
        print('Введено неправильное значение')

    execution_time = datetime.now() - start_time
    print(f'Время работы программы: {execution_time}')


if __name__ == '__main__':
    main()
