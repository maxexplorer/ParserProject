"""Сбор ИНН активных продавцов Wildberries и сохранение результата в Excel.

Скрипт проходит по диапазону seller_id, получает публичную статистику продавца,
проверяет активность по возрасту аккаунта и количеству товаров в продаже,
а затем сохраняет ссылку на продавца и его ИНН в XLSX-файл.
"""

import os
import time
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from requests import Response, Session

start_time = datetime.now()

USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36'
DEFAULT_429_PAUSE = 5
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RESULT_FILE_PATH = os.path.join(BASE_DIR, 'results', 'result_data_4_000_000.xlsx')
BATCH_SIZE = 50

SupplierData = dict[str, Any]
SellerResult = dict[str, str]


def get_api_headers(seller_id: int) -> dict[str, str]:
    """Сформировать HTTP-заголовки для запросов к публичным API Wildberries.

    Args:
        seller_id: Идентификатор продавца Wildberries.

    Returns:
        Словарь заголовков, имитирующий запросы из браузера.
    """
    return {
        'accept': '*/*',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'origin': 'https://www.wildberries.ru',
        'priority': 'u=1, i',
        'referer': f'https://www.wildberries.ru/seller/{seller_id}',
        'sec-ch-ua': '"Google Chrome";v="149", "Chromium";v="149", "Not)A;Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-site',
        'user-agent': USER_AGENT,
        'x-client-name': 'site',
    }


def sleep_if_429(response: Response, seller_id: int, request_name: str) -> bool:
    """Обработать ответ 429 Too Many Requests и выдержать паузу перед повтором.

    Args:
        response: Ответ сервера.
        seller_id: Идентификатор продавца, для которого выполнялся запрос.
        request_name: Человекочитаемое название запроса для логов.

    Returns:
        True, если сервер вернул 429 и была выполнена пауза, иначе False.
    """
    if response.status_code != 429:
        return False

    # Wildberries может вернуть рекомендуемую задержку в Retry-After.
    # Если заголовка нет или он некорректный, используем стандартную паузу.
    retry_after = response.headers.get('Retry-After')
    try:
        pause = int(retry_after) if retry_after else DEFAULT_429_PAUSE
    except ValueError:
        pause = DEFAULT_429_PAUSE

    print(f'Продавец {seller_id}: {request_name} вернул 429. Пауза {pause} секунд, затем повтор')
    time.sleep(pause)
    return True


def is_active_seller(years_on_wb: int, sale_item_quantity: int) -> bool:
    """Проверить, считается ли продавец активным по возрасту и числу товаров.

    Чем старше аккаунт продавца, тем выше минимальный порог товаров в продаже.

    Args:
        years_on_wb: Количество полных лет с даты регистрации продавца.
        sale_item_quantity: Количество товаров продавца в продаже.

    Returns:
        True, если продавец проходит порог активности.
    """
    if years_on_wb == 0:
        return sale_item_quantity >= 2000
    if years_on_wb == 1:
        return sale_item_quantity >= 4000
    if years_on_wb == 2:
        return sale_item_quantity >= 8000

    return sale_item_quantity >= 15000


def get_inn(session: Session, seller_id: int) -> str | None:
    """Получить ИНН продавца из публичного supplier-by-id endpoint.

    Args:
        session: Переиспользуемая HTTP-сессия.
        seller_id: Идентификатор продавца Wildberries.

    Returns:
        ИНН продавца, если его удалось получить, иначе None.
    """
    for _ in range(3):
        try:
            response = session.get(
                f'https://static-basket-01.wbbasket.ru/vol0/data/supplier-by-id/{seller_id}.json',
                headers=get_api_headers(seller_id),
                timeout=(3, 5)
            )

            if sleep_if_429(response, seller_id, 'запрос ИНН'):
                continue

            if response.status_code != 200:
                print(f'Продавец {seller_id}: статус ИНН {response.status_code}')
                return None

            json_data = response.json()
            inn = json_data.get('inn')
            return inn
        except Exception as ex:
            print(f'Ошибка при получении ИНН для {seller_id}: {ex}')
            time.sleep(DEFAULT_429_PAUSE)

    return None


def get_supplier_data(session: Session, seller_id: int) -> SupplierData | None:
    """Получить статистику продавца из suppliers-shipment API.

    Args:
        session: Переиспользуемая HTTP-сессия.
        seller_id: Идентификатор продавца Wildberries.

    Returns:
        JSON-данные продавца в виде словаря или None, если продавец не найден
        либо запрос завершился ошибкой.
    """
    params = {
        'curr': 'RUB',
    }

    while True:
        try:
            response = session.get(
                f'https://suppliers-shipment-2.wildberries.ru/api/v1/suppliers/{seller_id}',
                params=params,
                headers=get_api_headers(seller_id),
                timeout=(3, 5)
            )

            if sleep_if_429(response, seller_id, 'запрос статистики'):
                continue

            if response.status_code == 404:
                return None

            if response.status_code != 200:
                print(f'Продавец {seller_id}: статус статистики {response.status_code}')
                return None

            return response.json()
        except Exception as ex:
            print(f'Ошибка при получении статистики для {seller_id}: {ex}')
            return None


def get_active_seller_inn(session: Session, seller_id: int) -> str | None:
    """Получить ИНН продавца только если он проходит фильтр активности.

    Args:
        session: Переиспользуемая HTTP-сессия.
        seller_id: Идентификатор продавца Wildberries.

    Returns:
        ИНН активного продавца или None, если продавец неактивен либо данные
        получить не удалось.
    """
    json_data = get_supplier_data(session, seller_id)
    if not json_data:
        return None

    registration_date = json_data.get('registrationDate')
    sale_item_quantity = json_data.get('saleItemQuantity')

    if not registration_date or sale_item_quantity is None:
        return None

    # API возвращает дату регистрации в UTC-формате вроде 2024-01-31T12:34:56Z.
    reg_date = datetime.strptime(registration_date, '%Y-%m-%dT%H:%M:%SZ')
    years_on_wb = (datetime.now() - reg_date).days // 365

    if not is_active_seller(years_on_wb, sale_item_quantity):
        return None

    return get_inn(session, seller_id)


def process_sellers_range(start_id: int, end_id: int) -> None:
    """Обработать диапазон продавцов и сохранять найденные ИНН пачками.

    Args:
        start_id: Первый seller_id в диапазоне.
        end_id: Последний seller_id в диапазоне включительно.
    """
    result_list: list[SellerResult] = []

    with Session() as session:
        for seller_id in range(start_id, end_id + 1):
            try:
                inn = get_active_seller_inn(session, seller_id)

                if inn is None:
                    print(f'Обработан продавец ID: {seller_id}')
                    continue

                result_list.append(
                    {
                        'Ссылка': f'https://www.wildberries.ru/seller/{seller_id}',
                        'ИНН': inn
                    }
                )
                print(f'Обработан продавец ID: {seller_id}, inn: {inn}')
            except Exception as ex:
                print(f'https://www.wildberries.ru/seller/{seller_id}: {ex}')
                continue

            # Запись пачками снижает риск потери данных при долгом запуске.
            if len(result_list) >= BATCH_SIZE:
                save_excel(result_list)
                result_list.clear()

    if result_list:
        save_excel(result_list)


def save_excel(data: list[SellerResult]) -> None:
    """Добавить пачку результатов в XLSX-файл.

    Args:
        data: Список словарей с данными продавцов. Ключи первого элемента
            используются как заголовки при создании нового файла.
    """
    directory = os.path.dirname(RESULT_FILE_PATH)
    os.makedirs(directory, exist_ok=True)

    if not os.path.exists(RESULT_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sellers'
        excel_headers = list(data[0].keys())
        ws.append(excel_headers)
    else:
        wb = load_workbook(RESULT_FILE_PATH)
        ws = wb['Sellers'] if 'Sellers' in wb.sheetnames else wb.active

        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            excel_headers = list(data[0].keys())
            ws.append(excel_headers)
        else:
            excel_headers = []
            for cell in ws[1]:
                header = cell.value
                if isinstance(header, str) and header:
                    excel_headers.append(header)

    # Используем порядок заголовков из файла, чтобы новые строки совпадали
    # с уже существующей структурой таблицы.
    for row in data:
        ws.append([row.get(header) for header in excel_headers])

    wb.save(RESULT_FILE_PATH)
    print(f'Сохранено {len(data)} записей в {RESULT_FILE_PATH}')


def main() -> None:
    """Точка входа: запустить сбор по заданному диапазону seller_id."""
    start_id = 4_013_362
    end_id = 5_000_000

    process_sellers_range(start_id, end_id)

    execution_time = datetime.now() - start_time
    print('Сбор данных завершен.')
    print(f'Время выполнения: {execution_time}')


if __name__ == '__main__':
    main()
