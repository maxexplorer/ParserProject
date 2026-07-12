import os
import re
import time
from datetime import datetime
from typing import Any, TypeAlias

from openpyxl import Workbook, load_workbook
from requests import Response, Session

start_time = datetime.now()
SellerRow: TypeAlias = dict[str, Any]


class AuthExpiredError(Exception):
    """Сессия WB перестала подходить для запросов."""


USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36'
DEFAULT_429_PAUSE = 5
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_FILE_PATH = os.path.join(BASE_DIR, 'results', 'result_data_4_000_000.xlsx')
RESULT_FILE_PATH = os.path.join(BASE_DIR, 'results', 'result_data_4_000_000_with_products.xlsx')
START_ROW = 5402
BATCH_SIZE = 50
FILTERS_COOKIES = {
    'external-locale': 'ru',
    '_ga': 'GA1.1.1098996660.1758261326',
    '_ga_TXRZMJQDFE': 'GS2.1.s1759146135$o7$g0$t1759146135$j60$l0$h0',
    '_cp': '1',
    '__zzatw-wb': 'MDA0dBA=Fz2+aQ==',
    'cfidsw-wb': '/Cb1YoYAOLI/SCEhfiZq8bnsN+vkHtARk4gqgA18RS6OXNFdyhYUxbJ4hZrUcizLbBeETTj0/ZNVni2V+7QENlraGmsmO1zjyyprzCY8z9eb+NqimrZ2P+RKbDzDFrw5EOPOcl037573sXLDBCbz8vyJxMSTii0QuDR/',
    'routeb': '1779107226.277.2237.497709|fc3b37d75a18d923fd0e9c7589719997',
    'device_id': '2fabf07d-d2d6-4030-a541-d7db3985e3d1',
    'tours-city-id': '274286',
    'feedbacks_link_accepted': '1',
    '_wbauid': '2522710481783690074',
    'x_wbaas_token': '1.1000.91ceb65ffea74181a7d736137306f833.MTV8NDUuMTI5LjE0MS4xOTV8TW96aWxsYS81LjAgKFdpbmRvd3MgTlQgMTAuMDsgV2luNjQ7IHg2NCkgQXBwbGVXZWJLaXQvNTM3LjM2IChLSFRNTCwgbGlrZSBHZWNrbykgQ2hyb21lLzE1MC4wLjAuMCBTYWZhcmkvNTM3LjM2fDE3ODQwMjg0MzN8cmV1c2FibGV8MnxleUpvWVhOb0lqb2lJbjA9fDB8M3wxNzgzODk4ODMzfDE=.MEYCIQCrJgywud6osRc1VMCW7EELGAhRyVfeuA5dwA8NzSmwUAIhAP85yvlwLa6eWaEcazN6BDCQ9prke50oDVIPASulHs+m',
}


def get_catalog_headers(seller_id: int) -> dict[str, str]:
    """Вернуть HTTP-заголовки для эндпоинта каталога Wildberries."""

    return {
        'accept': '*/*',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'deviceid': 'site_d65c92c0ae19412c9cf011a89c998cf1',
        'priority': 'u=1, i',
        'referer': f'https://www.wildberries.ru/seller/{seller_id}',
        'sec-ch-ua': '"Not;A=Brand";v="8", "Chromium";v="150", "Google Chrome";v="150"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': USER_AGENT,
        'x-requested-with': 'XMLHttpRequest',
        'x-spa-version': '14.16.4',
    }


def sleep_if_429(response: Response, seller_id: int, request_name: str) -> bool:
    """Сделать паузу после rate limit и вернуть признак необходимости повтора."""

    if response.status_code != 429:
        return False

    retry_after = response.headers.get('Retry-After')
    try:
        pause = int(retry_after) if retry_after else DEFAULT_429_PAUSE
    except ValueError:
        pause = DEFAULT_429_PAUSE

    print(f'Продавец {seller_id}: {request_name} вернул 429. Пауза {pause} секунд, затем повтор')
    time.sleep(pause)
    return True


def extract_seller_id(seller_url: str) -> int | None:
    """Извлечь числовой ID продавца Wildberries из URL страницы продавца."""

    match = re.search(r'/seller/(\d+)', seller_url)
    if not match:
        return None

    return int(match.group(1))


def load_sellers(file_path: str) -> list[SellerRow]:
    """Загрузить строки продавцов из Excel и пропустить строки без ссылки."""

    wb = load_workbook(file_path)
    # В старых файлах может не быть листа Sellers, поэтому берем активный лист.
    ws = wb['Sellers'] if 'Sellers' in wb.sheetnames else wb.active
    headers = [cell.value for cell in ws[1]]

    sellers: list[SellerRow] = []
    for row in ws.iter_rows(min_row=START_ROW, values_only=True):
        row_data = dict(zip(headers, row))
        seller_url = row_data.get('Ссылка')
        if seller_url:
            sellers.append(row_data)

    return sellers


def has_products(session: Session, seller_id: int) -> bool:
    """Проверить, есть ли товары продавца Wildberries в поиске каталога."""

    params = {
        'ab_testing': 'false',
        'appType': '1',
        'curr': 'rub',
        'dest': '-1257786',
        'hide_dtype': '15',
        'hide_vflags': '4294967296',
        'lang': 'ru',
        'spp': '30',
        'supplier': str(seller_id),
    }

    while True:
        # Wildberries может временно ограничивать запросы; повторяем только при 429.
        response = session.get(
            'https://www.wildberries.ru/__internal/u-catalog/sellers/v8/filters',
            params=params,
            cookies=FILTERS_COOKIES,
            headers=get_catalog_headers(seller_id),
            timeout=(3, 5)
        )

        if sleep_if_429(response, seller_id, 'запрос каталога'):
            continue

        break

    if response.status_code == 498:
        raise AuthExpiredError(f'Продавец {seller_id}: 498, обнови cookies')

    if response.status_code != 200:
        print(f'Продавец {seller_id}: статус каталога {response.status_code}')
        return False

    json_data = response.json()
    total = json_data.get('data', {}).get('total')
    if not isinstance(total, int):
        print(f'Продавец {seller_id}: total не найден')
        return False

    return total > 0


def save_excel(data: list[SellerRow], file_path: str) -> None:
    """Сохранить строки продавцов в Excel, создав целевую папку при необходимости."""

    directory = os.path.dirname(file_path)
    os.makedirs(directory, exist_ok=True)

    if not data:
        return

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sellers'
        headers = list(data[0].keys())
        ws.append(headers)
    else:
        wb = load_workbook(file_path)
        ws = wb['Sellers'] if 'Sellers' in wb.sheetnames else wb.active
        headers = [cell.value for cell in ws[1] if isinstance(cell.value, str) and cell.value]

    for row in data:
        ws.append([row.get(header) for header in headers])

    wb.save(file_path)
    print(f'Сохранено {len(data)} записей в {file_path}')


def filter_sellers_with_products() -> None:
    """Отфильтровать продавцов по наличию товаров и записать итоговый файл."""

    if not os.path.exists(SOURCE_FILE_PATH):
        print(f'Файл не найден: {SOURCE_FILE_PATH}')
        return

    sellers = load_sellers(SOURCE_FILE_PATH)
    result_list: list[SellerRow] = []

    with Session() as session:
        for row_data in sellers:
            seller_url = row_data.get('Ссылка')
            if not isinstance(seller_url, str):
                continue

            seller_id = extract_seller_id(seller_url)
            if seller_id is None:
                print(f'Не удалось получить seller_id из ссылки: {seller_url}')
                continue

            try:
                if not has_products(session, seller_id):
                    print(f'Продавец {seller_id}: товаров нет')
                    continue

                result_list.append(row_data)
                if len(result_list) >= BATCH_SIZE:
                    save_excel(result_list, RESULT_FILE_PATH)
                    result_list.clear()
                print(f'Продавец {seller_id}: товары есть')
            except AuthExpiredError as ex:
                print(ex)
                if result_list:
                    save_excel(result_list, RESULT_FILE_PATH)
                    result_list.clear()
                print('Остановка: cookies/token устарели или не подходят.')
                return
            except Exception as ex:
                print(f'{seller_url}: {ex}')
                continue

    if result_list:
        save_excel(result_list, RESULT_FILE_PATH)


def main() -> None:
    """Запустить проверку наличия товаров и вывести время выполнения."""

    filter_sellers_with_products()

    execution_time = datetime.now() - start_time
    print('Проверка товаров завершена.')
    print(f'Время выполнения: {execution_time}')


if __name__ == '__main__':
    main()
