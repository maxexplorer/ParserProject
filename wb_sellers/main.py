import os
import time
from datetime import datetime
from random import randint
from typing import Any

from requests import Response, Session
from openpyxl import Workbook, load_workbook

start_time = datetime.now()
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36'
DEFAULT_429_PAUSE = 5
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RESULT_FILE_PATH = os.path.join(BASE_DIR, 'results', 'result_data_1.xlsx')
BATCH_SIZE = 50
FILTERS_COOKIES = {
    'external-locale': 'ru',
    '_ga': 'GA1.1.1098996660.1758261326',
    '_ga_TXRZMJQDFE': 'GS2.1.s1759146135$o7$g0$t1759146135$j60$l0$h0',
    '_cp': '1',
    '__zzatw-wb': 'MDA0dBA=Fz2+aQ==',
    'cfidsw-wb': '/Cb1YoYAOLI/SCEhfiZq8bnsN+vkHtARk4gqgA18RS6OXNFdyhYUxbJ4hZrUcizLbBeETTj0/ZNVni2V+7QENlraGmsmO1zjyyprzCY8z9eb+NqimrZ2P+RKbDzDFrw5EOPOcl037573sXLDBCbz8vyJxMSTii0QuDR/',
    'routeb': '1779107226.277.2237.497709|fc3b37d75a18d923fd0e9c7589719997',
    'device_id': '2fabf07d-d2d6-4030-a541-d7db3985e3d1',
    'tours-city-id': '274286',
    'feedbacks_link_accepted': '1',
    '_wbauid': '2522710481783690074',
    'x_wbaas_token': '1.1000.c5d279d3c77945eda3a58e1c8a304df5.MTV8NDUuMTI5LjE0MS4xOTV8TW96aWxsYS81LjAgKFdpbmRvd3MgTlQgMTAuMDsgV2luNjQ7IHg2NCkgQXBwbGVXZWJLaXQvNTM3LjM2IChLSFRNTCwgbGlrZSBHZWNrbykgQ2hyb21lLzE1MC4wLjAuMCBTYWZhcmkvNTM3LjM2fDE3ODQxMDI3Mzh8cmV1c2FibGV8MnxleUpvWVhOb0lqb2lJbjA9fDB8M3wxNzgzOTczMTM4fDE=.MEUCIBE9xkTq3bzri4cccqyyutBzKpGnJVFRLhpEHcm23GY+AiEA05JiXjau4rjSc9gYU+rpiZ3BITaB3GKs6cQnPqKYeAw=',
}

SupplierData = dict[str, Any]
SellerResult = dict[str, str]


def get_api_headers(seller_id: int) -> dict:
    return {
        'accept': '*/*',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'origin': 'https://www.wildberries.ru',
        'priority': 'u=1, i',
        'referer': f'https://www.wildberries.ru/seller/{seller_id}',
        'sec-ch-ua': '"Google Chrome";v="149", "Chromium";v="149", "Not)A;Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-site',
        'user-agent': USER_AGENT,
        'x-client-name': 'site',
    }


def get_catalog_headers(seller_id: int) -> dict:
    return {
        'accept': '*/*',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'deviceid': 'site_d65c92c0ae19412c9cf011a89c998cf1',
        'priority': 'u=1, i',
        'referer': f'https://www.wildberries.ru/seller/{seller_id}',
        'sec-ch-ua': '"Not;A=Brand";v="8", "Chromium";v="150", "Google Chrome";v="150"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': USER_AGENT,
        'x-requested-with': 'XMLHttpRequest',
        'x-spa-version': '14.16.4',
    }


def sleep_if_429(response: Response, seller_id: int, request_name: str) -> bool:
    if response.status_code != 429:
        return False

    retry_after = response.headers.get('Retry-After')
    try:
        pause = int(retry_after) if retry_after else DEFAULT_429_PAUSE
    except ValueError:
        pause = DEFAULT_429_PAUSE

    print(f'Продавец {seller_id}: {request_name} вернул 429. Пауза {pause} секунд, затем повтор')
    time.sleep(pause)
    return True


def is_active_seller(years_on_wb: int, sale_item_quantity: int) -> bool:
    if years_on_wb == 0:
        return sale_item_quantity >= 2000
    if years_on_wb == 1:
        return sale_item_quantity >= 4000
    if years_on_wb == 2:
        return sale_item_quantity >= 8000

    return sale_item_quantity >= 15000


def has_products(session: Session, seller_id: int) -> bool:
    params = {
        'ab_testing': 'false',
        'appType': '1',
        'curr': 'rub',
        'dest': '-1257786',
        'hide_dtype': '15',
        'hide_vflags': '4294967296',
        'lang': 'ru',
        'spp': '30',
        'supplier': str(seller_id),
    }

    while True:
        response = session.get(
            'https://www.wildberries.ru/__internal/u-catalog/sellers/v8/filters',
            params=params,
            cookies=FILTERS_COOKIES,
            headers=get_catalog_headers(seller_id),
            timeout=(3, 5)
        )

        if sleep_if_429(response, seller_id, 'запрос каталога'):
            continue

        break

    if response.status_code != 200:
        print(f'Продавец {seller_id}: статус каталога {response.status_code}')
        return False

    json_data = response.json()
    total = json_data.get('data', {}).get('total')
    if not isinstance(total, int):
        print(f'Продавец {seller_id}: total не найден')
        return False

    return total > 0


def get_inn(session: Session, seller_id: int) -> str | None:
    """
    Получает ИНН продавца по его ID через статичный JSON-эндпоинт WB.

    :param session: Активная сессия requests
    :param seller_id: ID продавца (строка)
    :return: ИНН как строка или None, если не найден
    """
    for _ in range(3):
        try:
            response = session.get(
                f'https://static-basket-01.wbbasket.ru/vol0/data/supplier-by-id/{seller_id}.json',
                headers=get_api_headers(seller_id),
                timeout=(3, 5)
            )

            if sleep_if_429(response, seller_id, 'запрос ИНН'):
                continue

            if response.status_code != 200:
                print(f'Продавец {seller_id}: статус ИНН {response.status_code}')
                return None

            json_data = response.json()
            inn = json_data.get('inn')
            return inn
        except Exception as ex:
            print(f'Ошибка при получении ИНН для {seller_id}: {ex}')
            time.sleep(DEFAULT_429_PAUSE)

    return None


def get_supplier_data(session: Session, seller_id: int) -> SupplierData | None:
    params = {
        'curr': 'RUB',
    }

    while True:
        try:
            response = session.get(
                f'https://suppliers-shipment-2.wildberries.ru/api/v1/suppliers/{seller_id}',
                params=params,
                headers=get_api_headers(seller_id),
                timeout=(3, 5)
            )

            if sleep_if_429(response, seller_id, 'запрос статистики'):
                continue

            if response.status_code == 404:
                return None

            if response.status_code != 200:
                print(f'Продавец {seller_id}: статус статистики {response.status_code}')
                return None

            return response.json()
        except Exception as ex:
            print(f'Ошибка при получении статистики для {seller_id}: {ex}')
            return None


def get_active_seller_inn(session: Session, seller_id: int) -> str | None:
    json_data = get_supplier_data(session, seller_id)
    if not json_data:
        return None

    registration_date = json_data.get('registrationDate')
    sale_item_quantity = json_data.get('saleItemQuantity')

    if not registration_date or sale_item_quantity is None:
        return None

    reg_date = datetime.strptime(registration_date, '%Y-%m-%dT%H:%M:%SZ')
    years_on_wb = (datetime.now() - reg_date).days // 365

    if not is_active_seller(years_on_wb, sale_item_quantity):
        return None

    return get_inn(session, seller_id)


def process_sellers_range(start_id: int, end_id: int) -> None:
    """
    Обрабатывает продавцов в заданном диапазоне ID.
    Проверяет, активен ли продавец, и сохраняет данные в Excel.

    :param start_id: Начальный ID (включительно)
    :param end_id: Конечный ID (включительно)
    """

    result_list = []

    with Session() as session:
        for seller_id in range(start_id, end_id + 1):
            try:
                if not has_products(session, seller_id):
                    print(f'Обработан продавец ID: {seller_id}')
                    continue

                inn = get_active_seller_inn(session, seller_id)

                if inn is None:
                    print(f'Обработан продавец ID: {seller_id}')
                    continue

                result_list.append(
                    {
                        'Ссылка': f'https://www.wildberries.ru/seller/{seller_id}',
                        'ИНН': inn
                    }
                )
                print(f'Обработан продавец ID: {seller_id}, inn: {inn}')

            except Exception as ex:
                print(f'https://www.wildberries.ru/seller/{seller_id}: {ex}')
                continue

            if len(result_list) >= BATCH_SIZE:
                save_excel(result_list)
                result_list.clear()

    if result_list:
        save_excel(result_list)


def save_excel(data: list[SellerResult]) -> None:
    """Добавить пачку результатов в XLSX-файл.

    Args:
        data: Список словарей с данными продавцов. Ключи первого элемента
            используются как заголовки при создании нового файла.
    """
    directory = os.path.dirname(RESULT_FILE_PATH)
    os.makedirs(directory, exist_ok=True)

    if not os.path.exists(RESULT_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sellers'
        excel_headers = list(data[0].keys())
        ws.append(excel_headers)
    else:
        wb = load_workbook(RESULT_FILE_PATH)
        ws = wb['Sellers'] if 'Sellers' in wb.sheetnames else wb.active

        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            excel_headers = list(data[0].keys())
            ws.append(excel_headers)
        else:
            excel_headers = []
            for cell in ws[1]:
                header = cell.value
                if isinstance(header, str) and header:
                    excel_headers.append(header)

    # Используем порядок заголовков из файла, чтобы новые строки совпадали
    # с уже существующей структурой таблицы.
    for row in data:
        ws.append([row.get(header) for header in excel_headers])

    wb.save(RESULT_FILE_PATH)
    print(f'Сохранено {len(data)} записей в {RESULT_FILE_PATH}')


def main() -> None:
    """
    Точка входа в программу. Запускает обработку продавцов в заданном диапазоне.
    """
    # Укажи нужный диапазон ID
    start_id = 1
    end_id = 1_700_000

    process_sellers_range(start_id, end_id)

    execution_time = datetime.now() - start_time
    print('Сбор данных завершен.')
    print(f'Время выполнения: {execution_time}')


if __name__ == '__main__':
    main()
