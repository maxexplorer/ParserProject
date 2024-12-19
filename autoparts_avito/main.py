import os
from datetime import datetime

from openpyxl import load_workbook, Workbook

start_time = datetime.now()

# Путь к файлу avito.xlsx
avito_file = 'avito/avito.xlsx'

# Загружаем данные из avito.xlsx
wb_avito = load_workbook(avito_file)
ws_avito = wb_avito.active

# Считываем артикулы и цены в словарь
avito_dict = {row[1].value: row[3].value for row in ws_avito.iter_rows(min_row=3, max_col=5)}


# Папка с файлами data
data_folder = 'data'

# Создаем новую книгу для записи найденных данных
new_wb = Workbook()
new_ws = new_wb.active
new_ws.append(['Артикул', 'Цена', 'Лист'])  # Заголовки

# Итерация по файлам в папке data
for file_name in os.listdir(data_folder):
    file_path = os.path.join(data_folder, file_name)

    # Проверяем, что это Excel-файл
    if file_name.endswith(('.xlsx', '.xlsm')):
        try:
            # Загружаем книгу
            workbook = load_workbook(file_path)
            sheet_names = workbook.sheetnames
        except Exception as e:
            raise f'Ошибка чтения файла: {e}'


        # Проходим по листам, начиная со второго
        for i, sheet_name in enumerate(sheet_names):
            if i == 0:  # Пропускаем первый лист
                continue

            sheet = workbook[sheet_name]

            # Получаем заголовки из первой строки (предположим, заголовки находятся в первой строке)
            headers = [cell.value for cell in sheet[2]]

            # Находим индексы колонок с нужными заголовками
            try:
                oem_column_index = headers.index("OEM")  # Ищем столбец "OEM"
                price_column_index = headers.index("Price")  # Ищем столбец "Price"
            except Exception as ex:
                print(f'{sheet}: {ex}')
                continue

            # Поиск и обновление цены
            for row in sheet.iter_rows(min_row=5):  # min_row=3 пропускает заголовок
                article_cell = row[oem_column_index].value  # Колонка с артикулом

                if article_cell in avito_dict:
                    # Обновляем цену
                    new_price = avito_dict[article_cell]
                    row[price_column_index].value = new_price  # Колонка с ценой

                    print(f'Обработано: {article_cell}: {new_price}')

                    # Записываем в новый файл
                    new_ws.append([article_cell, new_price, sheet_name])

        # Сохраняем изменения в исходный файл
        workbook.save(file_path)

if not os.path.exists('results'):
    os.mkdir('results')

# Сохраняем новый файл с найденными данными
new_wb.save('results/updated_prices.xlsx')

execution_time = datetime.now() - start_time
print('Сбор данных завершен!')
print(f'Время работы программы: {execution_time}')
