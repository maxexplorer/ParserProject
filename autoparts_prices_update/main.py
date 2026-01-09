import os
from datetime import datetime
from openpyxl import load_workbook, Workbook


def read_old_articles_data(old_articles_file):
    """Чтение данных из файла 'старые артикулы.xlsx'"""
    wb = load_workbook(old_articles_file)
    ws = wb.active

    articles_dict = {}

    for row in ws.iter_rows(min_row=2, max_col=2):
        name = row[0].value  # Наименование
        article = row[1].value  # Артикул

        if not name or not article:
            continue

        try:
            articles_dict[name.strip()] = str(article.replace('-', '')).strip()
        except Exception as ex:
            print(f'read_old_articles_data: {name} error: {ex}')

    return articles_dict


def read_new_articles_data(new_articles_file):
    """Чтение данных из файла 'старые артикулы.xlsx'"""
    wb = load_workbook(new_articles_file)
    ws = wb.active

    new_articles_dict = {}

    for row in ws.iter_rows(min_row=2, max_col=3):
        new_article = row[1].value  # Новый артикул
        old_article = row[2].value  # Старый артикул

        if not new_article or not old_article:
            continue

        try:
            new_articles_dict[old_article.strip()] = str(new_article).strip()
        except Exception as ex:
            print(f'read_new_articles_data: {new_article} error: {ex}')

    return new_articles_dict


def update_new_articles_file(new_articles_file, articles_dict):
    """Заполнение столбца 'Старый артикул' в файле 'новые артикулы.xlsx'"""
    wb = load_workbook(new_articles_file)
    ws = wb.active

    updated_count = 0

    for row in ws.iter_rows(min_row=2, max_col=3):
        name_cell = row[0]  # Наименование
        old_article_cell = row[2]  # Старый артикул

        if not name_cell.value:
            continue

        name = name_cell.value.strip()

        if name in articles_dict:
            old_article_cell.value = articles_dict[name]
            updated_count += 1

    wb.save(new_articles_file)
    print(f'Обновлено строк: {updated_count}')


def process_data_files(data_folder, articles_dict):
    """Обработка файлов из папки data и обновление цен"""
    # Итерация по файлам в папке data
    for file_name in os.listdir(data_folder):
        data_file_path = os.path.join(data_folder, file_name)

        print(f'Обрабатывается файл: {file_name}')

        # Проверяем, что это Excel-файл
        if file_name.endswith(('.xlsx', '.xlsm')):
            try:
                # Загружаем книгу
                workbook = load_workbook(data_file_path)
                sheet_names = workbook.sheetnames
            except Exception as e:
                raise f'Ошибка чтения файла: {e}'

            # Создаем новую книгу для записи результатов конкретного файла
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.append(['Артикул', 'Новый артикул', 'Лист'])  # Заголовки

            # Проходим по листам, начиная со второго
            for sheet_name in sheet_names:
                if not sheet_name.strip() == 'Для автомобилей-Кузов':
                    continue

                sheet = workbook[sheet_name]

                # Получаем заголовки из первой строки (предположим, заголовки находятся в первой строке)
                headers = [cell.value for cell in sheet[2]]

                # Находим индексы колонок с нужными заголовками
                try:
                    oem_column_index = headers.index('Номер детали OEM')  # Ищем столбец "OEM"
                except Exception as ex:
                    # print(f'{sheet}: {ex}')
                    continue

                # Поиск и обновление цены
                for row in sheet.iter_rows(min_row=5):  # min_row=5 пропускает заголовок
                    try:
                        article_cell = row[oem_column_index].value  # Колонка с артикулом
                        article_cell = article_cell.replace('-', '')
                    except Exception:
                        continue

                    if article_cell in articles_dict:
                        # Обновляем цену
                        new_article = articles_dict[article_cell]
                        row[oem_column_index].value = new_article  # Колонка с ценой

                        # Записываем в новый файл
                        new_ws.append([article_cell, new_article, sheet_name])

            # Сохраняем изменения в исходный файл
            workbook.save(data_file_path)

            # Сохраняем результаты
            save_results(new_wb=new_wb, file_name=file_name)


def save_results(new_wb, file_name):
    """Сохранение результатов в файл"""
    directory = 'results'
    os.makedirs(directory, exist_ok=True)

    result_file_path = f'results/updated_prices_{file_name}'

    new_wb.save(result_file_path)


def main():
    start_time = datetime.now()

    try:
        old_articles_file_path = 'data1/старые артикулы.xlsx'
        new_articles_file_path = 'data1/новые артикулы.xlsx'

        articles_dict = read_old_articles_data(old_articles_file_path)
        update_new_articles_file(new_articles_file_path, articles_dict)

        new_articles_dict = read_new_articles_data(new_articles_file_path)

        process_data_files(data_folder='data2', articles_dict=new_articles_dict)

    except Exception as ex:
        print(f'main: {ex}')
        input("Нажмите Enter, чтобы закрыть программу...")

    execution_time = datetime.now() - start_time
    print('Обработка завершена!')
    print(f'Время работы программы: {execution_time}')


if __name__ == '__main__':
    main()
