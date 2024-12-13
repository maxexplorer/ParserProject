import os
from openpyxl import load_workbook, Workbook

# Путь к файлу avito.xlsx
avito_file = "avito.xlsx"

# Загружаем данные из avito.xlsx
wb_avito = load_workbook(avito_file)
ws_avito = wb_avito.active

# Считываем артикулы и цены в словарь
avito_dict = {row[0].value: row[1].value for row in ws_avito.iter_rows(min_row=2, max_col=2)}

# Папка с файлами data
data_folder = "data"

# Создаем новую книгу для записи найденных данных
new_wb = Workbook()
new_ws = new_wb.active
new_ws.append(["Артикул", "Цена"])  # Заголовки

# Итерация по файлам в папке data
for file_name in os.listdir(data_folder):
    file_path = os.path.join(data_folder, file_name)

    # Проверяем, что это Excel-файл
    if file_name.endswith(('.xlsx', '.xlsm')):
        # Загружаем книгу
        workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames

        # Проходим по листам, начиная со второго
        for i, sheet_name in enumerate(sheet_names):
            if i == 0:  # Пропускаем первый лист
                continue

            sheet = workbook[sheet_name]

            # Поиск и обновление цены
            for row in sheet.iter_rows(min_row=2):  # min_row=2 пропускает заголовок
                article_cell = row[0]  # Первая колонка с артикулом
                price_cell = row[1]  # Вторая колонка с ценой

                if article_cell.value in avito_dict:
                    # Обновляем цену
                    new_price = avito_dict[article_cell.value]
                    price_cell.value = new_price

                    # Записываем в новый файл
                    new_ws.append([article_cell.value, new_price])

        # Сохраняем изменения в исходный файл
        workbook.save(file_path)

# Сохраняем новый файл с найденными данными
new_wb.save("updated_prices.xlsx")
