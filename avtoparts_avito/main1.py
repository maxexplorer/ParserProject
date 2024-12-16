import os
from datetime import datetime
from openpyxl import load_workbook, Workbook


def read_avito_data(avito_file):
    """Чтение данных из файла avito.xlsx"""
    wb_avito = load_workbook(avito_file)
    ws_avito = wb_avito.active
    avito_dict = {row[1].value: row[3].value for row in ws_avito.iter_rows(min_row=3, max_col=5)}
    return avito_dict


def process_data_files(data_folder, avito_dict, new_ws, new_wb):
    """Обработка файлов из папки data и обновление цен"""
    # Итерация по файлам в папке data
    for file_name in os.listdir(data_folder):
        data_file_path = os.path.join(data_folder, file_name)

        # Проверяем, что это Excel-файл
        if file_name.endswith(('.xlsx', '.xlsm')):
            try:
                # Загружаем книгу
                workbook = load_workbook(data_file_path)
                sheet_names = workbook.sheetnames
            except Exception as e:
                raise f'Ошибка чтения файла: {e}'

            # Проходим по листам, начиная со второго
            for i, sheet_name in enumerate(sheet_names):
                if i == 0:  # Пропускаем первый лист
                    continue

                sheet = workbook[sheet_name]

                # Получаем заголовки из первой строки (предположим, заголовки находятся в первой строке)
                headers = [cell.value for cell in sheet[2]]

                # Находим индексы колонок с нужными заголовками
                try:
                    oem_column_index = headers.index("OEM")  # Ищем столбец "OEM"
                    price_column_index = headers.index("Price")  # Ищем столбец "Price"
                except Exception as ex:
                    print(f'{sheet}: {ex}')
                    continue

                # Поиск и обновление цены
                for row in sheet.iter_rows(min_row=5):  # min_row=3 пропускает заголовок
                    article_cell = row[oem_column_index].value  # Колонка с артикулом

                    if article_cell in avito_dict:
                        # Обновляем цену
                        new_price = avito_dict[article_cell]
                        row[price_column_index].value = new_price  # Колонка с ценой

                        print(f'Обработано: {article_cell}: {new_price}')

                        # Записываем в новый файл
                        new_ws.append([article_cell, new_price, sheet_name])

            # Сохраняем изменения в исходный файл
            workbook.save(data_file_path)

            # Сохраняем результаты
            save_results(new_wb=new_wb, file_name=file_name)


def save_results(new_wb, file_name):
    """Сохранение результатов в файл"""
    data = 'results'

    if not os.path.exists(data):
        os.mkdir(data)

    result_file_path = f'results/updated_prices_{file_name}'

    new_wb.save(result_file_path)


def main():
    """Основная точка входа в программу"""
    start_time = datetime.now()

    # Путь к файлу avito.xlsx
    avito_file_path = 'avito/avito.xlsx'
    avito_dict = read_avito_data(avito_file_path)

    # Папка с файлами data
    data_folder = 'data'

    # Создаем новую книгу для записи найденных данных
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(['Артикул', 'Цена', 'Лист'])  # Заголовки

    # Обрабатываем файлы
    process_data_files(data_folder=data_folder, avito_dict=avito_dict, new_ws=new_ws, new_wb=new_wb)

    execution_time = datetime.now() - start_time
    print('Сбор данных завершен!')
    print(f'Время работы программы: {execution_time}')


if __name__ == "__main__":
    main()